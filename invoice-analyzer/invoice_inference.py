"""
Inferenza con modello LayoutLM (v1) per estrazione dati da fatture
Versione compatibile con Windows che utilizza le librerie Hugging Face
Versione corretta e migliorata, con rimozione di regex e sovrascrittura entità rule-based.
"""

import os
import re
import sys
import torch
import numpy as np
import pandas as pd
import json
import platform
import shutil
import warnings
import traceback
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
from transformers import LayoutLMTokenizer, LayoutLMForTokenClassification, LayoutLMConfig

os.environ["HF_HUB_OFFLINE"] = "1"
os.environ["TRANSFORMERS_OFFLINE"] = "1"

# Importazioni con gestione errori
try:
    from pdf2image import convert_from_path
    import pytesseract
except ImportError as e:
    print(f"Errore: Libreria mancante: {e}")
    print("Installa le librerie necessarie con: pip install pdf2image pytesseract transformers")
    sys.exit(1)

def _detect_path(binary_name, fallbacks=None):
    path = shutil.which(binary_name)
    if path:
        return path
    if fallbacks:
        for fb in fallbacks:
            if Path(fb).exists():
                return fb
    return binary_name

DEFAULT_CONFIG = {
    "POPPLER_PATH": str(Path(_detect_path("pdftoppm", ["/opt/homebrew/bin/pdftoppm", "/usr/bin/pdftoppm"])).parent),
    "TESSERACT_PATH": _detect_path("tesseract", ["/opt/homebrew/bin/tesseract", "/usr/bin/tesseract"]),
    "MODEL_DIR": "models/invoice_model",
    "OUTPUT_DIR": "output_data/results",
    "PDF_DPI": 300,
    "PDF_TIMEOUT": 120,
    "MAX_PAGES": 5,
    "MAX_SEQ_LENGTH": 512,
    "USE_LOCAL_MODEL": True
}

# Carica configurazione da file se esiste
def load_config():
    config = DEFAULT_CONFIG.copy()
    config_file = Path("invoice_config.json")
    
    if config_file.exists():
        try:
            with open(config_file, "r") as f:
                user_config = json.load(f)
                config.update(user_config)
            print(f"Configurazione caricata da {config_file}")
        except Exception as e:
            print(f"Errore nel caricamento della configurazione: {e}")
    
    for key in ["POPPLER_PATH", "TESSERACT_PATH"]:
        path = Path(config[key])
        if not path.exists():
            print(f"ATTENZIONE: Il percorso {key}={path} non esiste!")
    
    tesseract_path = Path(config["TESSERACT_PATH"])
    if tesseract_path.is_dir():
        tesseract_bin = "tesseract.exe" if platform.system() == "Windows" else "tesseract"
        config["TESSERACT_PATH"] = str(tesseract_path / tesseract_bin)
        print(f"Percorso Tesseract aggiornato a: {config['TESSERACT_PATH']}")
    
    return config

CONFIG = load_config()

LABELS = [
    "O", "B-VENDOR", "I-VENDOR", "B-CUSTOMER", "I-CUSTOMER", "B-DATE", "I-DATE",
    "B-TOTAL", "I-TOTAL", "B-ITEM", "I-ITEM", "B-QUANTITY", "I-QUANTITY",
    "B-PRICE", "I-PRICE", "B-INVOICE_NUMBER", "I-INVOICE_NUMBER"
]
id2label = {i: label for i, label in enumerate(LABELS)}
label2id = {label: i for i, label in enumerate(LABELS)}

##############################################################################################################################################################################################

def normalize_box(box, width, height):
    if width <= 0 or height <= 0:
        raise ValueError(f"Dimensioni dellimmagine non valide: {width}x{height}")
    x1 = min(max(0, box[0]), width - 1)
    y1 = min(max(0, box[1]), height - 1)
    x2 = min(max(x1 + 1, box[2]), width)
    y2 = min(max(y1 + 1, box[3]), height)
    normalized_box = [
        int(1000 * (x1 / width)),
        int(1000 * (y1 / height)),
        int(1000 * (x2 / width)),
        int(1000 * (y2 / height)),
    ]
    for i, coord in enumerate(normalized_box):
        if coord < 0 or coord > 1000:
            print(f"ATTENZIONE: Coordinata normalizzata fuori range: {coord}. Valore corretto.")
            normalized_box[i] = max(0, min(1000, coord))
    return normalized_box

def convert_pdf_to_images(pdf_path, dpi=None, poppler_path=None, max_pages=None):
    dpi = dpi or CONFIG["PDF_DPI"]
    poppler_path = poppler_path or CONFIG["POPPLER_PATH"]
    max_pages = max_pages or CONFIG["MAX_PAGES"]
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"Il file PDF non esiste: {pdf_path}")
    poppler_path = Path(poppler_path)
    if not poppler_path.exists():
        raise FileNotFoundError(f"Il percorso di Poppler non esiste: {poppler_path}")
    try:
        images = convert_from_path(
            str(pdf_path), dpi=dpi, poppler_path=str(poppler_path),
            first_page=1, last_page=max_pages, timeout=CONFIG["PDF_TIMEOUT"]
        )
        print(f"PDF convertito in {len(images)} immagini")
        return images
    except Exception as e:
        print(f"Errore nella conversione del PDF: {e}")
        if "timeout" in str(e).lower():
            try:
                print("Ritentativo con timeout esteso...")
                images = convert_from_path(
                    str(pdf_path), dpi=dpi, poppler_path=str(poppler_path),
                    first_page=1, last_page=max_pages, timeout=CONFIG["PDF_TIMEOUT"] * 2
                )
                print(f"PDF convertito in {len(images)} immagini con timeout esteso")
                return images
            except Exception as e2:
                print(f"Anche il tentativo con timeout esteso è fallito: {e2}")
        try:
            print("Tentativo di conversione pagina per pagina...")
            images = []
            for page_num in range(1, min(max_pages + 1, 6)):
                try:
                    page_images = convert_from_path(
                        str(pdf_path), dpi=dpi, poppler_path=str(poppler_path),
                        first_page=page_num, last_page=page_num, timeout=CONFIG["PDF_TIMEOUT"]
                    )
                    if page_images:
                        images.extend(page_images)
                        print(f"Pagina {page_num} convertita con successo")
                except Exception as e3:
                    print(f"Impossibile convertire la pagina {page_num}: {e3}")
            if images:
                print(f"PDF convertito in {len(images)} immagini con metodo alternativo")
                return images
            else:
                raise Exception("Nessuna pagina convertita con successo")
        except Exception as e4:
            print(f"Tutti i tentativi di conversione sono falliti: {e4}")
            raise

def process_image_with_ocr(image, tesseract_path=None):
    tesseract_path = tesseract_path or CONFIG["TESSERACT_PATH"]
    tesseract_path = Path(tesseract_path)
    if not tesseract_path.exists():
        raise FileNotFoundError(f"Il percorso di Tesseract non esiste: {tesseract_path}")
    pytesseract.pytesseract.tesseract_cmd = str(tesseract_path)
    try:
        print(f"Utilizzando Tesseract versione: {pytesseract.get_tesseract_version()}")
        ocr_result = pytesseract.image_to_data(
            image, output_type=pytesseract.Output.DICT, config="--psm 6 --oem 3 -l ita+eng"
        )
        words, boxes, word_indices = [], [], []
        for i in range(len(ocr_result["text"])):
            word = ocr_result["text"][i].strip()
            if word and ocr_result["conf"][i] > 30:
                x, y, w, h = ocr_result["left"][i], ocr_result["top"][i], ocr_result["width"][i], ocr_result["height"][i]
                if w > 1 and h > 1:
                    words.append(word)
                    boxes.append([x, y, x + w, y + h])
                    word_indices.append(i)
        print(f"Trovate {len(words)} parole nellimmagine con confidenza sufficiente")
        if not words:
            print("Nessuna parola trovata, tentativo con parametri OCR alternativi...")
            ocr_result = pytesseract.image_to_data(
                image, output_type=pytesseract.Output.DICT, config="--psm 4 --oem 3 -l ita+eng"
            )
            for i in range(len(ocr_result["text"])):
                word = ocr_result["text"][i].strip()
                if word:
                    x, y, w, h = ocr_result["left"][i], ocr_result["top"][i], ocr_result["width"][i], ocr_result["height"][i]
                    if w > 1 and h > 1:
                        words.append(word)
                        boxes.append([x, y, x + w, y + h])
                        word_indices.append(i)
            print(f"Trovate {len(words)} parole con metodo alternativo")
        if not words:
            raise ValueError("OCR non ha rilevato testo nellimmagine.")
        return words, boxes, word_indices, ocr_result
    except Exception as e:
        print(f"Errore nel processo OCR: {e}")
        traceback.print_exc()
        raise

def visualize_predictions(image, ocr_result, predictions, output_path):
    result_image = image.copy()
    draw = ImageDraw.Draw(result_image)
    colors = {
        "VENDOR": (255,0,0,150), "CUSTOMER": (0,0,255,150), "DATE": (0,255,0,150),
        "TOTAL": (128,0,128,150), "ITEM": (255,165,0,150), "QUANTITY": (0,255,255,150),
        "PRICE": (255,0,255,150), "INVOICE_NUMBER": (165,42,42,150), "O": (200,200,200,50)
    }
    legend_items = []
    entity_boxes = {}
    current_entity, current_text, current_box, current_indices = None, [], [float("inf"), float("inf"), 0, 0], []
    
    sorted_indices = sorted(predictions.keys(), key=lambda x: (ocr_result["top"][int(x)], ocr_result["left"][int(x)]))
    
    for word_idx in sorted_indices:
        i = int(word_idx)
        word = ocr_result["text"][i].strip()
        if not word: continue
        label = predictions[word_idx]
        x, y, w, h = ocr_result["left"][i], ocr_result["top"][i], ocr_result["width"][i], ocr_result["height"][i]

        if label.startswith("B-"):
            if current_entity and current_text:
                entity_type = current_entity
                if entity_type not in entity_boxes: entity_boxes[entity_type] = []
                avg_conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                entity_boxes[entity_type].append((" ".join(current_text), current_box, avg_conf))
                if entity_type not in [item[0] for item in legend_items]:
                    legend_items.append((entity_type, colors.get(entity_type, (0,0,0,150))))
            current_entity = label[2:]
            current_text = [word]
            current_box = [x, y, x + w, y + h]
            current_indices = [i]
        elif label.startswith("I-") and current_entity == label[2:]:
            current_text.append(word)
            current_box[0] = min(current_box[0], x)
            current_box[1] = min(current_box[1], y)
            current_box[2] = max(current_box[2], x + w)
            current_box[3] = max(current_box[3], y + h)
            current_indices.append(i)
        elif label == "O":
            if current_entity and current_text:
                entity_type = current_entity
                if entity_type not in entity_boxes: entity_boxes[entity_type] = []
                avg_conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                entity_boxes[entity_type].append((" ".join(current_text), current_box, avg_conf))
                if entity_type not in [item[0] for item in legend_items]:
                    legend_items.append((entity_type, colors.get(entity_type, (0,0,0,150))))
            current_entity, current_text, current_box, current_indices = None, [], [float("inf"), float("inf"), 0, 0], []
            
    if current_entity and current_text:
        entity_type = current_entity
        if entity_type not in entity_boxes: entity_boxes[entity_type] = []
        avg_conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
        entity_boxes[entity_type].append((" ".join(current_text), current_box, avg_conf))
        if entity_type not in [item[0] for item in legend_items]:
            legend_items.append((entity_type, colors.get(entity_type, (0,0,0,150))))

    for entity_type, entries in entity_boxes.items():
        color = colors.get(entity_type, (0,0,0,150))
        for text, box, confidence in entries:
            x1, y1, x2, y2 = box
            overlay = Image.new("RGBA", result_image.size, (0,0,0,0))
            draw_overlay = ImageDraw.Draw(overlay)
            draw_overlay.rectangle([x1, y1, x2, y2], fill=color)
            if result_image.mode != "RGBA": result_image = result_image.convert("RGBA")
            result_image = Image.alpha_composite(result_image, overlay)
            draw = ImageDraw.Draw(result_image)
            draw.rectangle([x1, y1, x2, y2], outline=color[:3], width=2)
            try:
                font = ImageFont.truetype("arial.ttf", 12)
            except IOError:
                font = ImageFont.load_default()
            label_text = f"{entity_type}: {text[:30]}{'...' if len(text) > 30 else ''}"

            try: label_w, label_h = draw.textbbox((0,0), label_text, font=font)[2:]
            except AttributeError: label_w, label_h = draw.textsize(label_text, font=font)
            draw.rectangle([x1, y1 - label_h - 4, x1 + label_w + 4, y1], fill=color[:3])
            draw.text((x1 + 2, y1 - label_h - 2), label_text, fill=(255,255,255), font=font)

    if legend_items:
        legend_x, legend_y, legend_width = 10, 10, 150
        legend_height = len(legend_items) * 20 + 30
        draw.rectangle([legend_x, legend_y, legend_x + legend_width, legend_y + legend_height], fill=(0,0,0,128), outline=(255,255,255))
        try: title_font = ImageFont.truetype("arial.ttf", 14)
        except IOError: title_font = ImageFont.load_default()
        draw.text((legend_x + 5, legend_y + 5), "Entità rilevate:", fill=(255,255,255), font=title_font)
        for i, (entity_type, color) in enumerate(legend_items):
            item_y = legend_y + 25 + i * 20
            draw.rectangle([legend_x + 5, item_y, legend_x + 15, item_y + 10], fill=color[:3], outline=(255,255,255))
            draw.text((legend_x + 20, item_y - 2), entity_type, fill=(255,255,255), font=font)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    result_image = result_image.convert("RGB")
    result_image.save(output_path)
    print(f"Visualizzazione salvata in: {output_path}")
    return result_image

def add_rule_based_entities(ocr_result, entities):
    """Versione migliorata che verifica meglio prima di sovrascrivere entità."""
    print("Esecuzione della verifica delle entità...")
    
    # Per DATE, INVOICE_NUMBER, TOTAL sovrascriviamo solo se la confidenza è alta
    
    # Cerca DATE con alta precisione
    dates_found_by_rule = []
    for i, text_token in enumerate(ocr_result["text"]):
        text = str(text_token).strip()
        if is_valid_date_format(text) and ocr_result["conf"][i] > 70:  # Alta confidenza
            print(f"Data verificata trovata: {text} (conf: {ocr_result['conf'][i]})")
            dates_found_by_rule.append({
                "text": text,
                "indices": [i],
                "confidence": ocr_result["conf"][i],
                "source": "rule"
            })
    
    # Verifica se sovrascrivere DATE
    if dates_found_by_rule:
        if "DATE" not in entities or not entities["DATE"]:
            # Se non c'è nessuna data rilevata dal modello, usa quelle dalle regole
            print(f"Nessuna data trovata dal modello. Usando: {[d['text'] for d in dates_found_by_rule]}")
            entities["DATE"] = dates_found_by_rule
        else:
            # Confronta confidenza delle date
            model_date_conf = max([d.get('confidence', 0) for d in entities["DATE"]])
            rule_date_conf = max([d.get('confidence', 0) for d in dates_found_by_rule])
            
            if rule_date_conf > model_date_conf + 10:  # Soglia per sovrascrittura
                print(f"Sovrascrittura date: regole più confidenti ({rule_date_conf} vs {model_date_conf})")
                entities["DATE"] = dates_found_by_rule
            else:
                print(f"Mantenute date del modello ({model_date_conf} vs {rule_date_conf})")
    
    # Cerca INVOICE_NUMBER con alta precisione
    invoice_numbers_found = []
    for i, text_token in enumerate(ocr_result["text"]):
        text = str(text_token).strip()
        if is_invoice_number_format(text) and ocr_result["conf"][i] > 75:
            # Verifica context per confermare che è un numero di fattura
            context_confirmed = False
            invoice_keywords = ["fattura", "invoice", "numero", "number", "n.", "facture", "rif", "ref"]
            
            # Controlla le parole circostanti (±3 token)
            for j in range(max(0, i-3), min(len(ocr_result["text"]), i+4)):
                if j != i and any(keyword in ocr_result["text"][j].lower() for keyword in invoice_keywords):
                    context_confirmed = True
                    print(f"Numero fattura verificato trovato: {text} (conf: {ocr_result['conf'][i]}, context: {ocr_result['text'][j]})")
                    break
            
            if context_confirmed or ocr_result["conf"][i] > 90:  # Alta confidenza o contesto confermato
                invoice_numbers_found.append({
                    "text": text,
                    "indices": [i],
                    "confidence": ocr_result["conf"][i] + (20 if context_confirmed else 0),  # Bonus per context
                    "source": "rule"
                })
    
    # Verifica se sovrascrivere INVOICE_NUMBER
    if invoice_numbers_found:
        if "INVOICE_NUMBER" not in entities or not entities["INVOICE_NUMBER"]:
            # Se non c'è nessun numero fattura rilevato dal modello, usa quelli dalle regole
            entities["INVOICE_NUMBER"] = invoice_numbers_found
            print(f"Numeri fattura trovati tramite regole: {[n['text'] for n in invoice_numbers_found]}")
        else:
            # Confronta confidenza
            model_conf = max([n.get('confidence', 0) for n in entities["INVOICE_NUMBER"]])
            rule_conf = max([n.get('confidence', 0) for n in invoice_numbers_found])
            
            if rule_conf > model_conf + 15:  # Soglia più alta per sovrascrittura
                print(f"Sovrascrittura numero fattura: regole più confidenti ({rule_conf} vs {model_conf})")
                entities["INVOICE_NUMBER"] = invoice_numbers_found
            else:
                print(f"Mantenuto numero fattura del modello ({model_conf} vs {rule_conf})")
    
    # Cerca TOTAL con analisi avanzata
    total_candidates = []
    
    # Prima cerca importi grandi
    amounts = []
    for i, text_token in enumerate(ocr_result["text"]):
        text = str(text_token).strip()
        if is_amount_format(text) and ocr_result["conf"][i] > 70:
            try:
                # Prova a convertire l'importo in un numero
                numeric_value = 0
                
                # Gestisci formato europeo (1.234,56)
                if "," in text and "." in text:
                    # Formato 1.234,56
                    clean_text = text.replace(".", "").replace(",", ".")
                    numeric_value = float(clean_text)
                elif "," in text:
                    # Formato 1234,56
                    clean_text = text.replace(",", ".")
                    numeric_value = float(clean_text)
                elif "." in text:
                    # Formato 1234.56
                    numeric_value = float(text)
                else:
                    # Solo cifre
                    numeric_value = float(text)
                
                # Cerca keyword TOTAL nelle vicinanze
                context_score = 0
                total_keywords = ["total", "totale", "importo", "amount", "sum", "montant", "eur", "euro", "€"]
                
                # Controlla nelle vicinanze (±3 token)
                for j in range(max(0, i-3), min(len(ocr_result["text"]), i+4)):
                    if j != i and any(keyword in ocr_result["text"][j].lower() for keyword in total_keywords):
                        context_score += 30
                        print(f"Importo {text} ({numeric_value}) trovato vicino a '{ocr_result['text'][j]}'")
                
                # Maggiore è l'importo, maggiore è la probabilità che sia un totale
                size_score = min(50, numeric_value / 1000)  # Max 50 punti per importi > €50.000
                
                # Posizione (i totali tendono ad essere nella parte inferiore)
                position_y = ocr_result["top"][i]
                doc_height = max(ocr_result["top"][k] + ocr_result["height"][k] for k in range(len(ocr_result["text"])))
                position_score = 0
                if doc_height > 0:
                    relative_position = position_y / doc_height
                    if relative_position > 0.7:  # Parte inferiore
                        position_score = 20
                
                # Score totale per questo importo
                total_score = ocr_result["conf"][i] + context_score + size_score + position_score
                
                amounts.append({
                    "text": text,
                    "value": numeric_value,
                    "confidence": ocr_result["conf"][i],
                    "context_score": context_score,
                    "size_score": size_score,
                    "position_score": position_score,
                    "total_score": total_score,
                    "index": i
                })
            except ValueError:
                continue
    
    # Ordina gli importi per punteggio totale
    amounts.sort(key=lambda x: x["total_score"], reverse=True)
    
    # Seleziona candidati di alta qualità per TOTAL
    for amount in amounts[:3]:  # Considera i primi 3 candidati
        if amount["total_score"] > 120:  # Alta qualità
            total_candidates.append({
                "text": amount["text"],
                "indices": [amount["index"]],
                "confidence": amount["confidence"],
                "total_score": amount["total_score"],
                "source": "amount_analysis"
            })
            print(f"Candidato TOTAL trovato: {amount['text']} (valore: {amount['value']}, score: {amount['total_score']})")
    
    # Verifica se sovrascrivere TOTAL
    if total_candidates:
        if "TOTAL" not in entities or not entities["TOTAL"]:
            # Se non c'è nessun totale rilevato dal modello, usa quelli dalla regola
            entities["TOTAL"] = [total_candidates[0]]
            print(f"TOTAL non trovato dal modello, usando analisi importi: {total_candidates[0]['text']}")
        else:
            # Confronta qualità
            model_total = entities["TOTAL"][0]
            rule_total = total_candidates[0]
            
            # Cerca di estrarre valori numerici per confronto
            model_value = 0
            rule_value = 0
            try:
                model_text = model_total['text'].replace(".", "").replace(",", ".")
                rule_text = rule_total['text'].replace(".", "").replace(",", ".")
                model_value = float(''.join(c for c in model_text if c.isdigit() or c == '.'))
                rule_value = float(''.join(c for c in rule_text if c.isdigit() or c == '.'))
            except:
                pass
            
            # Se l'importo trovato dalle regole è significativamente più grande o ha score molto più alto
            if (rule_value > model_value * 2) or (rule_total.get('total_score', 0) > model_total.get('total_score', 0) + 30):
                print(f"Sovrascrittura TOTAL: analisi importi più convincente ({rule_total['text']} vs {model_total['text']})")
                entities["TOTAL"] = [rule_total]
            else:
                print(f"Mantenuto TOTAL del modello: {model_total['text']}")
    
    # Verifica coerenza del VENDOR
    if "VENDOR" in entities and entities["VENDOR"]:
        vendor = entities["VENDOR"][0]['text'].lower()
        # Controlla se il vendor sembra essere una stringa generica o descrittiva
        generic_terms = ["purchase", "order", "number", "client", "invoice", "document"]
        if any(term in vendor for term in generic_terms):
            # Cerca un'alternativa migliore
            potential_vendors = []
            
            for i, text_token in enumerate(ocr_result["text"]):
                text = str(text_token).strip()
                if len(text) > 5 and ocr_result["conf"][i] > 80:
                    # Escludiamo testo che sembra una descrizione generica, numeri o date
                    if not any(term in text.lower() for term in generic_terms) and not is_amount_format(text) and not is_valid_date_format(text):
                        # Cerca parole chiave vendor nelle vicinanze
                        vendor_keywords = ["fornitore", "vendor", "supplier", "seller", "emittente", "company"]
                        context_score = 0
                        
                        # Controlla nelle vicinanze (±3 token)
                        for j in range(max(0, i-3), min(len(ocr_result["text"]), i+4)):
                            if j != i and any(keyword in ocr_result["text"][j].lower() for keyword in vendor_keywords):
                                context_score += 30
                                print(f"Candidato VENDOR '{text}' trovato vicino a '{ocr_result['text'][j]}'")
                        
                        # Calcola score per questo candidato
                        position_y = ocr_result["top"][i]
                        doc_height = max(ocr_result["top"][k] + ocr_result["height"][k] for k in range(len(ocr_result["text"])))
                        position_score = 0
                        if doc_height > 0:
                            relative_position = position_y / doc_height
                            if relative_position < 0.3:  # Parte superiore (i vendor sono tipicamente in alto)
                                position_score = 25
                        
                        score = ocr_result["conf"][i] + context_score + position_score
                        
                        potential_vendors.append({
                            "text": text,
                            "confidence": ocr_result["conf"][i],
                            "context_score": context_score,
                            "position_score": position_score,
                            "total_score": score,
                            "index": i
                        })
            
            # Ordina potenziali vendor
            potential_vendors.sort(key=lambda x: x["total_score"], reverse=True)
            
            if potential_vendors and potential_vendors[0]["total_score"] > 130:
                print(f"Sostituzione VENDOR generico con alternativa migliore: {potential_vendors[0]['text']}")
                entities["VENDOR"] = [{
                    "text": potential_vendors[0]["text"],
                    "indices": [potential_vendors[0]["index"]],
                    "confidence": potential_vendors[0]["confidence"],
                    "total_score": potential_vendors[0]["total_score"],
                    "source": "vendor_correction"
                }]
    
    # Stampa entità finali
    print("\nEntità finali (modello + regole verificate):")
    for entity_type, values in entities.items():
        print(f"  {entity_type}: {len(values)} occorrenze")
        for item in values:
            print(f"    - {item['text']} (confidenza: {item.get('confidence', 0):.1f})")
            
    return entities

def load_local_tokenizer(model_dir):
    print(f"Caricamento del tokenizer locale da {model_dir}...")
    model_dir = Path(model_dir)
    if not model_dir.exists():
        raise FileNotFoundError(f"Directory del modello non trovata: {model_dir}")
    try:
        try:
            from transformers import LayoutLMTokenizerFast
            print("Tentativo di caricamento del tokenizer fast...")
            tokenizer = LayoutLMTokenizerFast.from_pretrained(str(model_dir), local_files_only=True)
            print("Tokenizer LayoutLM Fast caricato con successo")
            return tokenizer
        except (ImportError, Exception) as e:
            print(f"Impossibile caricare il tokenizer fast: {e}")
        tokenizer = LayoutLMTokenizer.from_pretrained(str(model_dir), local_files_only=True)
        print("Tokenizer LayoutLM caricato con successo dalla directory locale")
        return tokenizer
    except Exception as e:
        print(f"Errore nel caricamento del tokenizer LayoutLM: {e}")
        from transformers import BertTokenizer # Fallback estremo
        print("Tentativo fallback con tokenizer BERT...")
        try:
            tokenizer = BertTokenizer.from_pretrained(str(model_dir), local_files_only=True)
            print("Tokenizer BERT caricato come fallback")
            return tokenizer
        except Exception as e_bert:
            print(f"Fallito anche caricamento tokenizer BERT: {e_bert}")
            raise

def enhance_extraction_with_direct_label_matching(structured_data, ocr_result):
    """
    Funzione di post-elaborazione che cerca direttamente coppie etichetta-valore 
    tipiche nelle fatture, indipendentemente dalle predizioni del modello.
    È un approccio generico che funziona con qualsiasi fattura.
    """
    print("\n=== Analisi diretta delle coppie etichetta-valore ===")
    
    # Importiamo re qui per essere sicuri
    import re
    
    # Dizionario di etichette da cercare per ogni tipo di campo (multilingua e generico)
    label_patterns = {
        "INVOICE_NUMBER": [
            "numero de facture", "numero fattura", "invoice number", "n. fattura", "n°", "facture n°", 
            "reference", "ref", "code facture", "codice fattura", "numéro de facture", "invoice no",
            "invoice #", "fattura n.", "n. doc", "numero documento", "factura", "rechnung nr",
            "rechnungsnummer", "invoice", "facture", "nummer", "number", "facture no", "facture n°",
            "no. facture", "número de factura", "nr. facture", "doc n", "document no", "no. document"
        ],
        "DATE": [
            "date de facture", "date facture", "invoice date", "data fattura", "data documento", 
            "data emissione", "date d'émission", "emission date", "date:", "data:",
            "date", "fecha", "datum", "data", "date de", "date du", "date of", "del",
            "emessa il", "issued on", "issue date", "fecha de emision", "ausstellungsdatum",
            "document date", "data doc", "date doc", "dated", "émis le", "in data"
        ],
        "CUSTOMER": [
            "adresse client", "client", "customer", "cliente", "destinatario", "adresse de facturation", 
            "billing address", "ship to", "indirizzo cliente", "adresse de livraison", "destinazione",
            "buyer", "purchaser", "a:", "to:", "recipient", "comprador", "kupfer", "kunde", "client:",
            "customer:", "billable to", "billed to", "client details", "customer details", "adresse du client",
            "bill to", "indirizzo di spedizione", "indirizzo di consegna", "delivery address", "deliver to",
            "shipped to", "destinataire", "cliente:", "kundenname", "cliente nombre", "addressee"
        ],
        "TOTAL": [
            "total à payer", "total", "totale", "amount due", "montant", "montant total", "somme", 
            "total ttc", "importo totale", "grand total", "totale complessivo", "total da pagare", 
            "net à payer", "netto a pagare", "montant à payer", "totale fattura", "total invoice",
            "total amount", "amount total", "total due", "total facture", "gesamtbetrag", "total factura",
            "total:", "somma", "importo", "importo dovuto", "importo netto", "netto", "importe total",
            "total payable", "total to pay", "to pay", "da pagare", "totale complessivo", "overall total",
            "sum total", "net total", "total net", "totale netto", "net amount", "grand total amount"
        ],
        "VENDOR": [
            "fornitore", "vendor", "supplier", "mittente", "seller", "venditore", "from", "da:", 
            "fournisseur", "emesso da", "emit", "issued by", "emittente", "azienda", "société",
            "ragione sociale", "ditta", "business", "company", "p.iva", "partita iva", "vat",
            "company name", "name", "from:", "sender", "emisor", "aussteller", "anbieter",
            "verkaufer", "absender", "issued from", "issued by", "emetteur", "nome azienda",
            "denominazione", "business name", "partita i.v.a.", "vat number", "c.f.", "nome"
        ]
    }
    
    # Converte tutto il testo OCR in minuscolo per il matching case-insensitive
    ocr_texts_lower = [text.lower() for text in ocr_result["text"]]
    
    # Risultati trovati
    found_entities = {}
    
    # Calcola l'altezza del documento per identificare posizioni relative
    doc_height = max([ocr_result["top"][i] + ocr_result["height"][i] for i in range(len(ocr_result["text"]))])
    doc_width = max([ocr_result["left"][i] + ocr_result["width"][i] for i in range(len(ocr_result["text"]))])
    
    # Per ogni tipo di campo, cerca le etichette corrispondenti
    for entity_type, labels in label_patterns.items():
        print(f"\nCercando etichette per {entity_type}:")
        
        for label in labels:
            # Cerca questa etichetta nel testo OCR
            for i, text_lower in enumerate(ocr_texts_lower):
                if label in text_lower:
                    print(f"  ✓ Trovata etichetta '{label}' in: '{ocr_result['text'][i]}'")
                    
                    # Esamina la posizione relativa nel documento
                    y_pos = ocr_result["top"][i] / doc_height if doc_height > 0 else 0.5
                    x_pos = ocr_result["left"][i] / doc_width if doc_width > 0 else 0.5
                    
                    # Bonus posizione in base al tipo di campo
                    position_score = 0
                    if entity_type in ["VENDOR", "CUSTOMER", "DATE", "INVOICE_NUMBER"] and y_pos < 0.5:
                        # Bonus per campi che solitamente appaiono nella parte alta
                        position_score = 30 * (1 - y_pos)
                    elif entity_type == "TOTAL" and y_pos > 0.5:
                        # Bonus per TOTAL nella parte bassa
                        position_score = 30 * y_pos
                    
                    # Cerca il valore nei token successivi all'etichetta
                    potential_values = []
                    
                    # Strategia di ricerca:
                    # 1. Prima guarda i token sulla stessa riga a destra dell'etichetta
                    # 2. Poi guarda i token sulla riga sotto
                    # 3. Infine cerca nei token successivi nell'ordine OCR
                    
                    label_right = ocr_result["left"][i] + ocr_result["width"][i]
                    label_top = ocr_result["top"][i] 
                    label_bottom = label_top + ocr_result["height"][i]
                    
                    # Cerca prima sulla stessa riga, a destra dell'etichetta
                    same_line_values = []
                    for j, text in enumerate(ocr_result["text"]):
                        if i == j or not text.strip():
                            continue
                            
                        j_left = ocr_result["left"][j]
                        j_top = ocr_result["top"][j]
                        j_bottom = j_top + ocr_result["height"][j]
                        
                        # Sulla stessa riga (o quasi)
                        is_same_line = abs(label_top - j_top) < max(ocr_result["height"][i], ocr_result["height"][j])
                        is_to_right = j_left > label_right
                        
                        if is_same_line and is_to_right:
                            # Verifica che il testo non sia vuoto
                            if text.strip():
                                proximity = 1.0 - (j_left - label_right) / (doc_width * 0.5)
                                proximity = max(0.1, min(1.0, proximity))
                                
                                # Verifica formato specifico in base al tipo di entità
                                format_bonus = 0
                                if entity_type == "INVOICE_NUMBER" and contains_digits_or_alphanumeric(text):
                                    format_bonus = 50
                                elif entity_type == "DATE" and is_valid_date_format(text):
                                    format_bonus = 80
                                elif entity_type == "TOTAL" and is_amount_format(text):
                                    format_bonus = 80
                                
                                total_confidence = ocr_result["conf"][j] + position_score + (proximity * 100) + format_bonus
                                
                                same_line_values.append({
                                    "text": text.strip(),
                                    "confidence": total_confidence,
                                    "index": j,
                                    "proximity": proximity,
                                    "is_same_line": True
                                })
                                
                                print(f"    → Valore su stessa riga: '{text}' "
                                      f"(conf: {ocr_result['conf'][j]}, pos: +{position_score}, "
                                      f"proximity: +{proximity*100:.1f}, formato: +{format_bonus}, "
                                      f"totale: {total_confidence:.1f})")
                    
                    # Poi cerca nella riga sotto (o nelle successive)
                    below_line_values = []
                    for j, text in enumerate(ocr_result["text"]):
                        if i == j or not text.strip():
                            continue
                            
                        j_left = ocr_result["left"][j]
                        j_top = ocr_result["top"][j]
                        
                        # Sotto la riga dell'etichetta
                        is_below = j_top > label_bottom
                        # In colonna con l'etichetta (± una tolleranza)
                        tolerance = min(doc_width * 0.2, 100)  # Tolleranza orizzontale
                        is_in_column = (
                            (j_left >= ocr_result["left"][i] - tolerance) and 
                            (j_left <= ocr_result["left"][i] + ocr_result["width"][i] + tolerance)
                        )
                        
                        if is_below and is_in_column:
                            # Distanza verticale (normalizzata)
                            vertical_distance = (j_top - label_bottom) / (doc_height * 0.2)
                            vertical_proximity = max(0.1, min(1.0, 1.0 - vertical_distance))
                            
                            # Verifica formato specifico in base al tipo di entità
                            format_bonus = 0
                            if entity_type == "INVOICE_NUMBER" and contains_digits_or_alphanumeric(text):
                                format_bonus = 50
                            elif entity_type == "DATE" and is_valid_date_format(text):
                                format_bonus = 80
                            elif entity_type == "TOTAL" and is_amount_format(text):
                                format_bonus = 80
                            elif entity_type in ["CUSTOMER", "VENDOR"] and len(text.split()) > 1:
                                format_bonus = 30  # Bonus per nomi complessi (più di una parola)
                            
                            total_confidence = ocr_result["conf"][j] + position_score + (vertical_proximity * 80) + format_bonus
                            
                            below_line_values.append({
                                "text": text.strip(),
                                "confidence": total_confidence,
                                "index": j,
                                "proximity": vertical_proximity,
                                "is_below": True
                            })
                            
                            print(f"    → Valore sotto: '{text}' "
                                  f"(conf: {ocr_result['conf'][j]}, pos: +{position_score}, "
                                  f"proximity: +{vertical_proximity*80:.1f}, formato: +{format_bonus}, "
                                  f"totale: {total_confidence:.1f})")
                                
                    # Se abbiamo trovato valori sulla stessa riga, privilegiamoli
                    if same_line_values:
                        potential_values = same_line_values
                    # Altrimenti, se abbiamo trovato valori sotto, usiamo quelli
                    elif below_line_values:
                        potential_values = below_line_values
                    # Se non abbiamo trovato nulla, cerchiamo anche nei token successivi (generici)
                    else:
                        # Cerca nei token successivi (max 5)
                        for offset in range(1, 6):
                            next_idx = i + offset
                            if next_idx < len(ocr_result["text"]):
                                next_text = ocr_result["text"][next_idx].strip()
                                
                                if next_text:
                                    proximity_bonus = 100 - (offset * 15)  # 85, 70, 55, 40, 25 punti bonus
                                    
                                    # Verifica formato specifico in base al tipo di entità
                                    format_bonus = 0
                                    if entity_type == "INVOICE_NUMBER" and contains_digits_or_alphanumeric(next_text):
                                        format_bonus = 50
                                    elif entity_type == "DATE" and is_valid_date_format(next_text):
                                        format_bonus = 80
                                    elif entity_type == "TOTAL" and is_amount_format(next_text):
                                        format_bonus = 80
                                    
                                    total_confidence = ocr_result["conf"][next_idx] + position_score + proximity_bonus + format_bonus
                                    
                                    potential_values.append({
                                        "text": next_text,
                                        "confidence": total_confidence,
                                        "index": next_idx,
                                        "proximity": proximity_bonus / 100,
                                        "is_sequence": True
                                    })
                                    
                                    print(f"    → Valore sequenziale: '{next_text}' "
                                          f"(conf: {ocr_result['conf'][next_idx]}, pos: +{position_score}, "
                                          f"proximity: +{proximity_bonus}, formato: +{format_bonus}, "
                                          f"totale: {total_confidence:.1f})")
                    
                    # Se abbiamo trovato potenziali valori, seleziona il migliore
                    if potential_values:
                        best_value = max(potential_values, key=lambda x: x["confidence"])
                        
                        if entity_type not in found_entities:
                            found_entities[entity_type] = []
                        
                        found_entities[entity_type].append({
                            "text": best_value["text"],
                            "confidence": best_value["confidence"],
                            "indices": [best_value["index"]],
                            "total_score": best_value["confidence"],
                            "source": "direct_label_matching",
                            "label": label,
                            "position_score": position_score
                        })
                        
                        print(f"  ✓ Valore migliore per '{label}': '{best_value['text']}' "
                              f"(confidenza: {best_value['confidence']})")
    
    # Per ogni tipo di entità, seleziona il valore migliore tra quelli trovati
    direct_extraction_results = {}
    
    for entity_type, values in found_entities.items():
        if values:
            # Ordina per confidenza
            sorted_values = sorted(values, key=lambda x: x["confidence"], reverse=True)
            best_match = sorted_values[0]
            
            print(f"\nMiglior match per {entity_type}: '{best_match['text']}' "
                  f"(etichetta: '{best_match['label']}', confidenza: {best_match['confidence']})")
            
            direct_extraction_results[entity_type] = [best_match]
    
    # Se abbiamo trovato valori diretti, aggiornare i risultati strutturati
    enhanced_data = structured_data.copy()
    
    for entity_type, direct_values in direct_extraction_results.items():
        if direct_values:
            best_direct = direct_values[0]
            
            # Confronta con i risultati esistenti, se presenti
            if entity_type in enhanced_data and enhanced_data[entity_type]:
                best_existing = max(enhanced_data[entity_type], 
                                    key=lambda x: x.get("total_score", x.get("confidence", 0)))
                
                existing_score = best_existing.get("total_score", best_existing.get("confidence", 0))
                direct_score = best_direct.get("total_score", best_direct.get("confidence", 0))
                
                # Se il risultato diretto è più confidante o il formato è più appropriato
                better_format = False
                if entity_type == "TOTAL" and is_amount_format(best_direct['text']) and not is_amount_format(best_existing['text']):
                    better_format = True
                elif entity_type == "DATE" and is_valid_date_format(best_direct['text']) and not is_valid_date_format(best_existing['text']):
                    better_format = True
                
                if direct_score > existing_score or better_format:
                    print(f"⚠️ Sostituito {entity_type} '{best_existing['text']}' "
                          f"(score: {existing_score}) con '{best_direct['text']}' "
                          f"(score: {direct_score}) trovato con matching diretto")
                    enhanced_data[entity_type] = [best_direct]
            else:
                # Se non c'erano risultati esistenti, aggiungi quelli diretti
                enhanced_data[entity_type] = direct_values
                print(f"✅ Aggiunto nuovo {entity_type}: '{best_direct['text']}'")
    
    print("\n=== Fine dell'analisi diretta delle coppie etichetta-valore ===")
    return enhanced_data

def contains_digits_or_alphanumeric(text):
    """Verifica se una stringa contiene numeri o è alfanumerica con almeno un numero."""
    return any(c.isdigit() for c in text) or (any(c.isalpha() for c in text) and any(c.isdigit() for c in text))

##############################################################################################################################################################################################

def extract_structured_data(ocr_result, predictions):
    """
    Estrae dati strutturati dalle predizioni del modello con validazione e analisi contestuale avanzata.
    
    Args:
        ocr_result: Risultato OCR con testo e posizioni
        predictions: Predizioni dal modello per ogni parola
        
    Returns:
        Dizionario con entità estratte, filtrate e ordinate per rilevanza
    """
    entities = {}
    current_entity, current_words, current_indices = None, [], []
    print("Predizioni totali:", len(predictions))
    print("Labels trovate:", set(predictions.values()))
    
    # Ordina le parole per posizione (dall'alto a sinistra verso il basso a destra)
    word_indices_sorted = sorted(predictions.keys(), key=lambda x: (ocr_result["top"][int(x)], ocr_result["left"][int(x)]))
    
    # Primo passaggio: raccogli tutte le entità dal modello
    for word_idx_str in word_indices_sorted:
        i = int(word_idx_str)
        word = ocr_result["text"][i].strip()
        label = predictions[word_idx_str]
        
        if label != "O":
            print(f"Predizione: {word} -> {label} (conf: {ocr_result['conf'][i]})")

        if label.startswith("B-"):
            # Salva l'entità precedente se esiste
            if current_entity and current_words:
                if current_entity not in entities: entities[current_entity] = []
                conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                entities[current_entity].append({
                    "text": " ".join(current_words), 
                    "indices": current_indices, 
                    "confidence": conf,
                    "box": {
                        "x1": min(ocr_result["left"][j] for j in current_indices),
                        "y1": min(ocr_result["top"][j] for j in current_indices),
                        "x2": max(ocr_result["left"][j] + ocr_result["width"][j] for j in current_indices),
                        "y2": max(ocr_result["top"][j] + ocr_result["height"][j] for j in current_indices)
                    }
                })
            # Inizia una nuova entità
            current_entity = label[2:]
            current_words = [word]
            current_indices = [i]
        elif label.startswith("I-"):
            # Controlla se questa è una continuazione dell'entità corrente
            if not current_entity or current_entity != label[2:]:
                # Se non è una continuazione valida, inizia una nuova entità
                if current_entity and current_words:
                    if current_entity not in entities: entities[current_entity] = []
                    conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                    entities[current_entity].append({
                        "text": " ".join(current_words), 
                        "indices": current_indices, 
                        "confidence": conf,
                        "box": {
                            "x1": min(ocr_result["left"][j] for j in current_indices),
                            "y1": min(ocr_result["top"][j] for j in current_indices),
                            "x2": max(ocr_result["left"][j] + ocr_result["width"][j] for j in current_indices),
                            "y2": max(ocr_result["top"][j] + ocr_result["height"][j] for j in current_indices)
                        }
                    })
                current_entity = label[2:] # Start new entity as if it was B-
                current_words = [word]
                current_indices = [i]
            else:
                # Continua l'entità corrente
                current_words.append(word)
                current_indices.append(i)
        elif label == "O":
            # Fine dell'entità corrente
            if current_entity and current_words:
                if current_entity not in entities: entities[current_entity] = []
                conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                entities[current_entity].append({
                    "text": " ".join(current_words), 
                    "indices": current_indices, 
                    "confidence": conf,
                    "box": {
                        "x1": min(ocr_result["left"][j] for j in current_indices),
                        "y1": min(ocr_result["top"][j] for j in current_indices),
                        "x2": max(ocr_result["left"][j] + ocr_result["width"][j] for j in current_indices),
                        "y2": max(ocr_result["top"][j] + ocr_result["height"][j] for j in current_indices)
                    }
                })
            current_entity, current_words, current_indices = None, [], []
    
    # Assicurati di salvare l'ultima entità se esiste
    if current_entity and current_words:
        if current_entity not in entities: entities[current_entity] = []
        conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
        entities[current_entity].append({
            "text": " ".join(current_words), 
            "indices": current_indices, 
            "confidence": conf,
            "box": {
                "x1": min(ocr_result["left"][j] for j in current_indices),
                "y1": min(ocr_result["top"][j] for j in current_indices),
                "x2": max(ocr_result["left"][j] + ocr_result["width"][j] for j in current_indices),
                "y2": max(ocr_result["top"][j] + ocr_result["height"][j] for j in current_indices)
            }
        })

    # Stampa entità trovate solo dal modello
    print("\nEntità (solo modello) trovate:")
    for entity_type, values in entities.items():
        print(f"  {entity_type}: {len(values)} occorrenze")
        for item in values:
            print(f"    - {item['text']} (confidenza: {item.get('confidence', 0):.1f})")
    
    # Cerca di fondere entità vicine dello stesso tipo
    merged_entities = {}
    for entity_type, items in entities.items():
        if entity_type not in merged_entities:
            merged_entities[entity_type] = []
        
        # Ordina gli item per posizione (dall'alto a sinistra)
        sorted_items = sorted(items, key=lambda x: (x['box']['y1'], x['box']['x1']))
        
        # Primo tentativo di fusione per vicinanza spaziale
        merged_items = []
        current_group = []
        
        for item in sorted_items:
            if not current_group:
                current_group.append(item)
                continue
                
            last_item = current_group[-1]
            # Calcola distanza e altezza media
            avg_height = (last_item['box']['y2'] - last_item['box']['y1'] + 
                          item['box']['y2'] - item['box']['y1']) / 2
                          
            # Se le y sono simili (stessa linea)
            y_diff = abs(item['box']['y1'] - last_item['box']['y1'])
            x_diff = item['box']['x1'] - last_item['box']['x2']
            
            # Sulla stessa linea e vicini orizzontalmente (entro 2 volte l'altezza media)
            if y_diff < avg_height * 0.5 and x_diff < avg_height * 2 and x_diff > -avg_height * 2:
                current_group.append(item)
            else:
                # Salva il gruppo corrente e inizia uno nuovo
                if len(current_group) > 0:
                    merged_text = " ".join([i['text'] for i in current_group])
                    all_indices = [idx for item in current_group for idx in item['indices']]
                    avg_conf = sum(ocr_result["conf"][j] for j in all_indices) / len(all_indices) if all_indices else 0
                    merged_items.append({
                        "text": merged_text,
                        "indices": all_indices,
                        "confidence": avg_conf,
                        "box": {
                            "x1": min(item['box']['x1'] for item in current_group),
                            "y1": min(item['box']['y1'] for item in current_group),
                            "x2": max(item['box']['x2'] for item in current_group),
                            "y2": max(item['box']['y2'] for item in current_group)
                        }
                    })
                current_group = [item]
        
        # Aggiungi l'ultimo gruppo
        if current_group:
            merged_text = " ".join([i['text'] for i in current_group])
            all_indices = [idx for item in current_group for idx in item['indices']]
            avg_conf = sum(ocr_result["conf"][j] for j in all_indices) / len(all_indices) if all_indices else 0
            merged_items.append({
                "text": merged_text,
                "indices": all_indices,
                "confidence": avg_conf,
                "box": {
                    "x1": min(item['box']['x1'] for item in current_group),
                    "y1": min(item['box']['y1'] for item in current_group),
                    "x2": max(item['box']['x2'] for item in current_group),
                    "y2": max(item['box']['y2'] for item in current_group)
                }
            })
        
        # Aggiungi gli item fusi o gli originali se non è stato possibile fonderli
        if merged_items:
            merged_entities[entity_type].extend(merged_items)
        else:
            merged_entities[entity_type].extend(items)
    
    # Stampa entità dopo la fusione spaziale
    print("\nEntità dopo il tentativo di fusione spaziale:")
    for entity_type, values in merged_entities.items():
        print(f"  {entity_type}: {len(values)} occorrenze")
        for item in values:
            print(f"    - {item['text']} (confidenza: {item.get('confidence', 0):.1f})")
    
    # Valuta e filtra le entità
    validated_entities = {}
    for entity_type, values in merged_entities.items():
        validated_values = []
        
        # Ordina prima per confidenza, così da processare i candidati migliori prima
        values_sorted = sorted(values, key=lambda x: x.get('confidence', 0), reverse=True)
        
        for item in values_sorted:
            is_valid = validate_entity(entity_type, item['text'], item.get('confidence', 0))
            
            if is_valid:
                # Calcola score di contesto
                context_score = evaluate_context_coherence(entity_type, item, ocr_result)
                item['context_score'] = context_score
                
                # Calcola score semantico
                semantic_score = evaluate_entity_semantics(entity_type, item['text'])
                item['semantic_score'] = semantic_score
                
                # Bonus per VENDOR con più parole (nomi aziendali)
                if entity_type == "VENDOR" and len(item['text'].split()) > 1:
                    item['name_complexity_bonus'] = 25 * len(item['text'].split())  # Bonus per complessità del nome
                    print(f"  Bonus complessità per '{item['text']}': +{item['name_complexity_bonus']}")
                else:
                    item['name_complexity_bonus'] = 0
                
                # Bonus speciale per entità TOTAL che sono chiaramente numeri
                if entity_type == "TOTAL" and is_amount_format(item['text']):
                    item['numeric_format_bonus'] = 30  # Bonus per formato numerico
                    print(f"  Bonus formato numerico per TOTAL '{item['text']}': +30")
                else:
                    item['numeric_format_bonus'] = 0
                
                # Score totale: confidenza base + contesto + semantica + bonus specifici
                item['total_score'] = (item.get('confidence', 0) + 
                                      context_score + 
                                      semantic_score +
                                      item.get('name_complexity_bonus', 0) +
                                      item.get('numeric_format_bonus', 0))
                
                validated_values.append(item)
            else:
                print(f"  Scartata entità {entity_type} '{item['text']}' (non valida)")
        
        if validated_values:
            # Ordina per punteggio totale (confidenza + contesto + semantica + bonus)
            validated_values.sort(key=lambda x: x.get('total_score', 0), reverse=True)
            validated_entities[entity_type] = validated_values
    
    # Stampa entità validate
    print("\nEntità validate (con bonus):")
    for entity_type, values in validated_entities.items():
        print(f"  {entity_type}: {len(values)} occorrenze")
        for item in values:
            print(f"    - {item['text']} (confidenza: {item.get('confidence', 0):.1f}, "
                  f"contesto: {item.get('context_score', 0):.1f}, "
                  f"semantica: {item.get('semantic_score', 0):.1f}, "
                  f"bonus nome: {item.get('name_complexity_bonus', 0):.1f}, "
                  f"bonus numerico: {item.get('numeric_format_bonus', 0):.1f}, "
                  f"totale: {item.get('total_score', 0):.1f})")
    
    # Ricerca diretta di importi candidati per TOTAL (in caso il modello non ne abbia trovati validi)
    candidate_totals = find_candidate_totals(ocr_result)
    
    # Se abbiamo trovato importi candidati, consideriamoli
    if candidate_totals:
        print("\nImporti candidati a TOTAL trovati tramite analisi diretta:")
        for i, candidate in enumerate(candidate_totals[:5]):  # Mostra i top 5
            print(f"  {i+1}. '{candidate['text']}' (score: {candidate['score']:.1f}, valore: {candidate['value']:.2f})")
    
    # Selezione finale delle migliori entità (filtrate per soglia)
    final_entities = {}
    min_thresholds = {
        "VENDOR": 75,  # Manteniamo una soglia relativamente bassa per vendor
        "CUSTOMER": 80,
        "DATE": 75,
        "TOTAL": 75,  # Abbassiamo anche per TOTAL
        "INVOICE_NUMBER": 80
    }
    
    for entity_type, values in validated_entities.items():
        threshold = min_thresholds.get(entity_type, 75)
        filtered_values = [item for item in values if item.get('total_score', 0) >= threshold]
        
        if filtered_values:
            if entity_type in ["TOTAL", "DATE", "INVOICE_NUMBER"]:
                final_entities[entity_type] = [filtered_values[0]]  # Solo il migliore
            else:
                # Per VENDOR e CUSTOMER, preferiamo nomi complessi
                if entity_type == "VENDOR":
                    complex_names = [item for item in filtered_values if len(item['text'].split()) > 1]
                    if complex_names:
                        final_entities[entity_type] = [max(complex_names, 
                                                         key=lambda x: x.get('total_score', 0))]
                    else:
                        final_entities[entity_type] = [filtered_values[0]]
                else:
                    final_entities[entity_type] = [filtered_values[0]]  # Il migliore in assoluto
    
    # Integrazione di importi candidati per TOTAL se non ne abbiamo trovati validi
    if "TOTAL" not in final_entities or not final_entities["TOTAL"]:
        # Usa i candidati trovati con analisi diretta
        if candidate_totals and candidate_totals[0]["score"] > 120:  # Soglia alta per evitare falsi positivi
            print(f"\nAggiunto TOTAL '{candidate_totals[0]['text']}' da analisi diretta degli importi")
            final_entities["TOTAL"] = [{
                "text": candidate_totals[0]["text"],
                "indices": [candidate_totals[0]["index"]],
                "confidence": candidate_totals[0]["confidence"],
                "total_score": candidate_totals[0]["score"],
                "source": "amount_analysis"
            }]
    
    # Stampa entità finali
    print("\nEntità finali (post-selezione):")
    for entity_type, values in final_entities.items():
        print(f"  {entity_type}: {len(values)} occorrenze")
        for item in values:
            print(f"    - {item['text']} (score totale: {item.get('total_score', 0):.1f})")
    
    # Aggiungi regole specifiche se necessario (mantenuto per compatibilità)
    final_entities = add_rule_based_entities(ocr_result, final_entities)
    
    return final_entities

def apply_positional_heuristics(entities, ocr_result):
    """Applica euristiche basate sulla posizione per migliorare le entità rilevate."""
    # Ottieni l'altezza del documento
    doc_height = max([ocr_result["top"][i] + ocr_result["height"][i] for i in range(len(ocr_result["text"]))])
    
    # Calcola punteggi posizionali per ogni entità
    positional_scores = {}
    
    for entity_type in entities:
        for item_idx, item in enumerate(entities[entity_type]):
            # Calcola posizione verticale relativa
            if 'box' in item:
                y_position = item['box']['y1']
            elif 'indices' in item:
                indices = item['indices']
                y_position = min([ocr_result["top"][int(idx)] for idx in indices]) if indices else 0
            else:
                continue
                
            relative_position = y_position / doc_height if doc_height > 0 else 0.5
            
            # Calcola bonus basati su posizione
            position_bonus = 0
            
            # Vendor e Customer tendono ad essere nella parte superiore
            if entity_type in ["VENDOR", "CUSTOMER"] and relative_position < 0.4:
                position_bonus = 50 * (1 - relative_position * 2)  # Più è in alto, maggiore è il bonus
            
            # DATE e INVOICE_NUMBER tendono ad essere nella parte superiore/media
            elif entity_type in ["DATE", "INVOICE_NUMBER"] and relative_position < 0.5:
                position_bonus = 30 * (1 - relative_position)
            
            # TOTAL tende ad essere nella parte inferiore
            elif entity_type == "TOTAL" and relative_position > 0.6:
                position_bonus = 50 * (relative_position - 0.6) * 2.5  # Più è in basso, maggiore è il bonus
            
            # Penalizzazione per posizioni improbabili
            elif entity_type == "TOTAL" and relative_position < 0.5:
                position_bonus = -40  # Penalizza totali nella parte superiore
            elif entity_type in ["VENDOR", "CUSTOMER"] and relative_position > 0.7:
                position_bonus = -30  # Penalizza vendor/customer nella parte inferiore
            
            # Aggiorna il punteggio dell'entità
            if position_bonus != 0:
                item['position_bonus'] = position_bonus
                if 'total_score' in item:
                    item['total_score'] += position_bonus
                else:
                    item['total_score'] = item.get('confidence', 0) + position_bonus
                
                print(f"  Bonus posizionale per {entity_type} '{item['text']}': {position_bonus:+.1f} (pos={relative_position:.2f})")
    
    return entities

def validate_entity(entity_type, text, confidence):
    """Funzione di validazione migliorata per filtrare entità non valide."""
    
    if confidence < 70:  # Filtro base sulla confidenza
        return False
        
    # Controlli specifici migliorati per tipo di entità
    if entity_type == "VENDOR":
        # Escludiamo parole singole molto comuni o generiche
        lower_text = text.lower()
        generic_words = ["client", "customer", "vendor", "fornitore", "address", "adresse", 
                        "purchase", "order", "number", "shipping", "delivery", "reference"]
        
        # Controlla se contiene solo parole generiche
        words = lower_text.split()
        if all(word in generic_words for word in words):
            print(f"  Rifiutato {entity_type} '{text}' - termini generici")
            return False
        
        # Se è una singola parola generica, rifiuta
        if lower_text in generic_words and len(text.split()) == 1:
            print(f"  Rifiutato {entity_type} '{text}' - termine generico singolo")
            return False
            
        # Un fornitore non dovrebbe essere solo un numero
        if is_amount_format(text):
            print(f"  Rifiutato {entity_type} '{text}' - sembra un importo")
            return False
        
        # I vendor tendono ad avere nomi più complessi di 1-2 parole
        if len(text.split()) > 1:
            return True
            
    elif entity_type == "TOTAL":
        # Un totale deve essere un formato numerico valido
        return is_amount_format(text)
            
    elif entity_type == "DATE":
        # Una data deve essere in formato valido 
        return is_valid_date_format(text)
            
    elif entity_type == "INVOICE_NUMBER":
        # Un numero di fattura deve seguire un formato specifico
        return is_invoice_number_format(text)
            
    return True

def evaluate_context_coherence(entity_type, entity_data, ocr_result):
    """Versione migliorata che valuta la coerenza contestuale delle entità."""
    text = entity_data['text']
    
    # Definizione delle parole chiave per ogni tipo di entità - ESTESE
    context_keywords = {
        "CUSTOMER": ["cliente", "customer", "client", "destinatario", "ship to", "buyer", "acquirente", 
                    "destinazione", "consegna a", "dest.", "address", "adresse", "indirizzo", "cliente:", 
                    "acheteur", "deliver to", "billing", "recapito", "fatturazione"],
        "VENDOR": ["fornitore", "vendor", "supplier", "mittente", "seller", "venditore", "from", "da:", 
                  "fournisseur", "emesso da", "emit", "issued by", "emittente", "azienda", "società",
                  "ragione sociale", "ditta", "business", "company", "p.iva", "partita iva", "vat"],
        "TOTAL": ["total", "totale", "amount", "importo", "sum", "eur", "€", "somma", "montant", 
                 "total ttc", "total ht", "netto", "imponibile", "pagare", "à payer", "da pagare",
                 "importo totale", "totale complessivo", "grand total", "overall", "complessivo"],
        "DATE": ["date", "data", "emessa il", "issued", "invoice date", "data fattura", "data documento",
                "date facture", "émis le", "date d'émission", "date:", "data:", "del:", "in data", 
                "eseguito il", "performed on", "emission", "emissione", "issued on"],
        "INVOICE_NUMBER": ["invoice", "number", "fattura", "numero", "n.", "rif.", "ref", "référence",
                         "numéro", "facture n°", "fattura n.", "codice", "documento", "document",
                         "invoice no", "n. fattura", "num.", "n°", "No:", "riferimento", "ID"]
    }
    
    # Estrai gli indici delle parole nel documento
    indices = entity_data.get('indices', [])
    
    # Se non ci sono indici, non possiamo fare l'analisi contestuale
    if not indices:
        return 0
    
    # Costruisci una finestra contestuale per cercare le parole chiave
    context_words = []
    context_window = 8  # Guarda fino a 8 token prima/dopo
    
    for idx in indices:
        idx = int(idx)
        # Cerca prima e dopo l'entità
        for offset in range(-context_window, context_window + 1):
            check_idx = idx + offset
            if 0 <= check_idx < len(ocr_result["text"]) and check_idx not in indices:
                context_word = ocr_result["text"][check_idx].lower().strip()
                if context_word:
                    context_words.append({
                        "text": context_word,
                        "idx": check_idx,
                        "offset": offset  # Quanto è lontano dall'entità
                    })
    
    # Calcola il punteggio in base alla presenza di parole chiave nel contesto
    bonus_score = 0
    found_keywords = set()
    
    if entity_type in context_keywords:
        for keyword in context_keywords[entity_type]:
            for ctx_word in context_words:
                if keyword in ctx_word["text"] and keyword not in found_keywords:
                    # Il bonus è più alto se la parola chiave è più vicina
                    proximity_factor = max(0.5, 1 - (abs(ctx_word["offset"]) / context_window))
                    keyword_bonus = 15 * proximity_factor
                    
                    # Bonus extra se è una corrispondenza esatta
                    if keyword == ctx_word["text"]:
                        keyword_bonus *= 1.5
                    
                    bonus_score += keyword_bonus
                    found_keywords.add(keyword)
                    
                    print(f"  Bonus contestuale per {entity_type} '{text}': keyword '{keyword}' "
                          f"(distanza: {ctx_word['offset']}, bonus: +{keyword_bonus:.1f})")
    
    return bonus_score

##############################################################################################################################################################################################

def evaluate_entity_semantics(entity_type, text):
    """Valuta la coerenza semantica dell'entità in base al suo contenuto."""
    
    # Liste di termini che NON dovrebbero essere identificati come specifiche entità
    non_vendor_terms = ["purchase", "order", "number", "shipping", "delivery", "reference", 
                       "ref", "no.", "code", "vat", "fiscal", "payment", "terms"]
    
    non_customer_terms = ["ship", "to", "delivery", "destination", "deliver", "consignee"]
    
    non_total_terms = ["subtotal", "tax", "vat", "payment", "delivery", "shipping", 
                      "discount", "description", "code"]
    
    semantic_score = 0
    
    if entity_type == "VENDOR":
        # Verifica se il testo contiene termini che NON dovrebbero essere un venditore
        lower_text = text.lower()
        matched_terms = [term for term in non_vendor_terms if term.lower() in lower_text]
        if matched_terms:
            penalty = min(50, 20 * len(matched_terms))  # Penalizzazione proporzionale al numero di termini trovati
            semantic_score -= penalty
            print(f"  Penalizzazione semantica per VENDOR '{text}': -{penalty} (contiene termini generici: {', '.join(matched_terms)})")
    
    elif entity_type == "CUSTOMER":
        lower_text = text.lower()
        matched_terms = [term for term in non_customer_terms if term.lower() in lower_text]
        if matched_terms:
            penalty = min(30, 15 * len(matched_terms))
            semantic_score -= penalty
            print(f"  Penalizzazione semantica per CUSTOMER '{text}': -{penalty} (contiene termini generici)")
            
    elif entity_type == "TOTAL":
        lower_text = text.lower()
        matched_terms = [term for term in non_total_terms if term.lower() in lower_text]
        if matched_terms:
            penalty = min(30, 15 * len(matched_terms))
            semantic_score -= penalty
            print(f"  Penalizzazione semantica per TOTAL '{text}': -{penalty} (contiene termini descrittivi)")
            
    return semantic_score

def parse_amount_to_float(text):
    """Converte un importo testuale in float, gestendo vari formati."""
    # Copia di sicurezza del testo originale
    original_text = text
    
    # Rimuovi simboli di valuta e spazi
    for symbol in ["€", "$", "£", "¥", "EUR", "USD", "GBP", "JPY"]:
        text = text.replace(symbol, "").strip()
    
    text = text.replace(" ", "")
    
    # Estrai solo la parte che sembra un numero con eventuali separatori
    digits_and_separators = ""
    for char in text:
        if char.isdigit() or char in ",.":
            digits_and_separators += char
    
    text = digits_and_separators
    if not text:
        raise ValueError(f"Nessun formato numerico trovato in: {original_text}")
    
    try:
        # Gestisci formato europeo (1.234,56)
        if "," in text and ("." in text or any(c.isdigit() for c in text.split(",")[0])):
            # Formato europeo: 1.234,56 o 1234,56
            text = text.replace(".", "")  # Rimuovi separatori migliaia
            text = text.replace(",", ".")  # Converti separatore decimale
            return float(text)
        
        # Gestisci formato US (1,234.56)
        elif "." in text:
            # Formato US: 1,234.56 o 1234.56
            text = text.replace(",", "")  # Rimuovi separatori migliaia
            return float(text)
        
        # Gestisci numeri con virgola senza punto (es. 1234,56)
        elif "," in text:
            text = text.replace(",", ".")
            return float(text)
        
        # Gestisci interi 
        elif text.isdigit():
            return float(text)
        
        # Non riconosciuto, solleva eccezione
        raise ValueError(f"Formato importo non riconosciuto: {original_text}")
    except:
        raise ValueError(f"Impossibile convertire in float: {original_text}")

def find_candidate_totals(ocr_result):
    """Cerca importi candidati che potrebbero essere totali, con maggiore enfasi sulla posizione e parole chiave."""
    amounts = []
    
    # Trova parole chiave che indicano totali
    total_keywords = ["total", "totale", "importo", "amount", "sum", "montant", "somma", "complessivo", "finale", 
                     "da pagare", "payable", "€", "eur", "euro"]
    total_keyword_positions = []
    
    # Prima passa: identifica posizione delle parole chiave relative ai totali
    for i, text in enumerate(ocr_result["text"]):
        if text.strip().lower() in total_keywords or any(keyword in text.strip().lower() for keyword in total_keywords):
            y_pos = ocr_result["top"][i]
            total_keyword_positions.append({
                "text": text.strip(),
                "index": i,
                "y_position": y_pos
            })
            print(f"Parola chiave 'total' trovata: '{text}' alla posizione y={y_pos}")
    
    # Seconda passa: cerca importi nel documento
    for i, text in enumerate(ocr_result["text"]):
        if not text.strip():
            continue
            
        # Cerca testo che potrebbe contenere un importo
        if any(c.isdigit() for c in text):
            confidence = ocr_result["conf"][i]
            value_text = text.strip()
            
            # Prova a convertire in numero
            try:
                numeric_value = parse_amount_to_float(value_text)
                
                # Ignora importi troppo piccoli (< 10) che probabilmente sono quantità o altro
                if numeric_value < 10:
                    continue
                
                # Calcola punteggio base
                score = confidence
                
                # NUOVO: calcola distanza dalle parole chiave "total"
                min_distance = float('inf')
                closest_keyword = None
                
                for keyword_info in total_keyword_positions:
                    # Distanza verticale (più importante)
                    y_distance = abs(ocr_result["top"][i] - keyword_info["y_position"])
                    
                    # Considera anche la distanza orizzontale per parole sulla stessa riga
                    x_distance = 0
                    if abs(y_distance) < 20:  # Se sulla stessa riga o quasi
                        x_distance = abs(ocr_result["left"][i] - ocr_result["left"][keyword_info["index"]])
                        # Se l'importo è a destra della keyword, è più probabile che sia un totale
                        if ocr_result["left"][i] > ocr_result["left"][keyword_info["index"]]:
                            x_distance *= 0.5  # Dimezza la distanza per favorire questa configurazione
                    
                    # Distanza totale (pesa di più la componente verticale)
                    distance = y_distance + x_distance * 0.3
                    
                    if distance < min_distance:
                        min_distance = distance
                        closest_keyword = keyword_info
                
                # Bonus per vicinanza a parole chiave
                if closest_keyword and min_distance < 200:  # Se c'è una keyword vicina
                    proximity_bonus = max(0, 100 - min_distance * 0.5)
                    score += proximity_bonus
                    print(f"  Bonus per vicinanza a '{closest_keyword['text']}': +{proximity_bonus:.1f} (dist={min_distance:.1f})")
                
                # Verifica posizione verticale (importi totali tendono ad essere nella parte inferiore)
                y_position = ocr_result["top"][i]
                height = ocr_result["height"][i]
                image_height = max([ocr_result["top"][k] + ocr_result["height"][k] for k in range(len(ocr_result["text"]))])
                relative_position = y_position / image_height if image_height > 0 else 0.5
                
                # Bonus per importi grandi (probabilmente totali)
                if numeric_value > 100:
                    size_bonus = min(50, numeric_value / 100)  # Fino a 50 punti
                    score += size_bonus
                    print(f"  Bonus per importo grande: {value_text} -> +{size_bonus:.1f}")
                
                # Bonus MAGGIORE per posizione in basso nel documento
                if relative_position > 0.65:  # Ultimo 35% del documento
                    position_bonus = (relative_position - 0.65) * 200  # Fino a 70 punti
                    score += position_bonus
                    print(f"  Bonus per posizione in basso: +{position_bonus:.1f} (pos={relative_position:.2f})")
                
                amounts.append({
                    "text": value_text,
                    "value": numeric_value,
                    "confidence": confidence,
                    "score": score,
                    "index": i,
                    "position": relative_position,
                    "y_position": y_position
                })
            except ValueError:
                # Non è un importo numerico valido
                pass
    
    # Trova l'importo più grande
    if amounts:
        max_value = max(amounts, key=lambda x: x["value"])["value"]
        
        # Aggiorna i punteggi: bonus molto forte per l'importo più grande nella parte inferiore del documento
        for amount in amounts:
            # Se è un importo grande e nella parte inferiore, molto probabilmente è un totale
            if amount["value"] > max_value * 0.7 and amount["position"] > 0.65:
                amount["score"] += 80
                print(f"  Bonus speciale per importo grande in basso: {amount['text']} -> +80")
    
    # Ordina per punteggio
    amounts.sort(key=lambda x: x["score"], reverse=True)
    return amounts

def calculate_text_similarity(text1, text2):
    """
    Calcola la somiglianza tra due stringhe di testo.
    Ritorna un valore tra 0 e 1, dove 1 significa testi identici.
    """
    # Normalizza i testi (rimuovi spazi aggiuntivi, converti in minuscolo)
    text1 = text1.lower().strip()
    text2 = text2.lower().strip()
    
    # Se i testi sono identici, ritorna 1
    if text1 == text2:
        return 1.0
    
    # Calcola la distanza di Levenshtein (numero di modifiche necessarie per trasformare text1 in text2)
    def levenshtein_distance(s1, s2):
        if len(s1) < len(s2):
            return levenshtein_distance(s2, s1)
        
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    # Calcola la distanza
    distance = levenshtein_distance(text1, text2)
    
    # Calcola la similarità come 1 - (distanza / lunghezza massima)
    max_length = max(len(text1), len(text2))
    if max_length == 0:
        return 0.0
    
    similarity = 1.0 - (distance / max_length)
    return similarity

def extract_total_with_enhanced_detection(ocr_result):
    """
    Funzione specializzata per estrarre il valore TOTAL con maggiore precisione,
    cercando pattern tipici di importi totali nelle fatture.
    Approccio completamente generico.
    """
    print("\n=== Ricerca avanzata del campo TOTAL ===")
    
    # Liste di frasi che potrebbero precedere il TOTAL (multilingue)
    total_labels = [
        "total à payer", "total", "totale", "amount due", "montant", "montant total", 
        "total ttc", "importo totale", "grand total", "total da pagare", "somme",
        "net à payer", "netto a pagare", "importo", "amount", "sum", "total facture",
        "total amount", "total due", "total:", "gesamtbetrag", "total factura",
        "somma", "importo dovuto", "importo netto", "netto", "importe total",
        "à payer", "da pagare", "totale complessivo", "grand total amount"
    ]
    
    # Cerca queste etichette nel testo OCR (case insensitive)
    label_positions = []
    for i, text in enumerate(ocr_result["text"]):
        text_lower = text.lower().strip()
        
        for label in total_labels:
            if label in text_lower:
                label_positions.append({
                    "index": i,
                    "text": text,
                    "label": label,
                    "y_position": ocr_result["top"][i],
                    "x_position": ocr_result["left"][i]
                })
                print(f"Trovata etichetta TOTAL: '{text}' alla posizione ({ocr_result['left'][i]}, {ocr_result['top'][i]})")
    
    if not label_positions:
        print("Nessuna etichetta TOTAL trovata, cercando importi con caratteristiche da totale...")
        return find_totals_by_characteristics(ocr_result)
    
    # Per ogni etichetta trovata, cerca importi nelle vicinanze
    # (preferibilmente sotto o a destra)
    potential_totals = []
    
    # Calcola l'altezza del documento per posizionamento relativo
    doc_height = max([ocr_result["top"][i] + ocr_result["height"][i] for i in range(len(ocr_result["text"]))])
    doc_width = max([ocr_result["left"][i] + ocr_result["width"][i] for i in range(len(ocr_result["text"]))])
    
    for label_pos in label_positions:
        idx = label_pos["index"]
        label_x = label_pos["x_position"]
        label_y = label_pos["y_position"]
        label_width = ocr_result["width"][idx]
        label_height = ocr_result["height"][idx]
        
        # Cerca in tutte le parole che potrebbero essere importi
        for i, text in enumerate(ocr_result["text"]):
            # Salta se è la stessa parola dell'etichetta
            if i == idx:
                continue
                
            # Verifica se è un potenziale importo
            if is_amount_format(text):
                word_x = ocr_result["left"][i]
                word_y = ocr_result["top"][i]
                word_height = ocr_result["height"][i]
                
                # Calcola la posizione relativa rispetto all'etichetta
                x_diff = word_x - (label_x + label_width)
                y_diff = word_y - label_y
                
                # Verifica se è sulla stessa riga (o quasi)
                is_same_line = abs(y_diff) < max(label_height, word_height) * 1.5
                
                # Calcola la distanza euclidea
                distance = (x_diff**2 + y_diff**2) ** 0.5
                
                # Posizione relativa nel documento (i totali tendono ad essere in basso)
                relative_y = word_y / doc_height if doc_height > 0 else 0.5
                
                # Preferisci importi che sono sulla stessa riga a destra o sotto l'etichetta
                position_bonus = 0
                if is_same_line and x_diff > 0:  # Stessa riga, a destra
                    position_bonus += 80
                elif y_diff > 0 and y_diff < label_height * 3:  # Sotto, ma non troppo lontano
                    position_bonus += 60
                
                # Bonus per posizione in basso nel documento (dove spesso si trova il totale)
                if relative_y > 0.6:  # Parte inferiore del documento
                    position_bonus += 40 * (relative_y - 0.6) * 2.5  # Max 40 punti
                
                # Confidenza OCR di base
                base_confidence = ocr_result["conf"][i]
                
                # Bonus per importi con formato specifico (es. con decimali)
                format_bonus = 0
                has_decimals = "," in text or "." in text
                if has_decimals:  # Ha decimali
                    format_bonus += 30
                
                # Penalità per distanza
                distance_penalty = min(100, distance / 3)
                
                # Calcola punteggio totale
                total_score = base_confidence + position_bonus + format_bonus - distance_penalty
                
                # Tenta di convertire in valore numerico
                try:
                    numeric_value = parse_amount_to_float(text)
                    
                    # Bonus per importi grandi (più probabilmente totali)
                    if numeric_value > 100:
                        size_bonus = min(50, numeric_value / 200)
                        total_score += size_bonus
                    
                    potential_totals.append({
                        "text": text,
                        "value": numeric_value,
                        "confidence": base_confidence,
                        "distance": distance,
                        "position_bonus": position_bonus,
                        "format_bonus": format_bonus,
                        "distance_penalty": distance_penalty,
                        "total_score": total_score,
                        "index": i,
                        "related_label": label_pos["text"],
                        "is_same_line": is_same_line,
                        "has_decimals": has_decimals,
                        "relative_y": relative_y
                    })
                    
                    print(f"Potenziale TOTAL: '{text}' (valore: {numeric_value}, "
                          f"distanza: {distance:.1f}, score: {total_score:.1f})")
                except ValueError:
                    # Non è un importo valido
                    pass
    
    # Trova l'importo più grande
    max_value = 0
    if potential_totals:
        max_value = max(potential_totals, key=lambda x: x["value"])["value"]
        
        # Bonus per gli importi più grandi (verosimilmente totali)
        for total in potential_totals:
            # Se l'importo è almeno il 80% del massimo, probabilmente è un totale
            if total["value"] > max_value * 0.8:
                total["total_score"] += 40
                print(f"  Bonus per importo vicino al massimo: {total['text']} -> +40")
            # Penalizza gli importi molto piccoli rispetto al massimo
            elif total["value"] < max_value * 0.5:
                total["total_score"] -= 30
                print(f"  Penalità per importo piccolo: {total['text']} -> -30")
    
    # Seleziona il miglior candidato
    if potential_totals:
        # Ordina per punteggio
        potential_totals.sort(key=lambda x: x["total_score"], reverse=True)
        best_total = potential_totals[0]
        
        print(f"Miglior candidato TOTAL: '{best_total['text']}' (score: {best_total['total_score']:.1f})")
        
        # Crea il risultato nel formato atteso
        result = {
            "text": best_total["text"],
            "indices": [best_total["index"]],
            "confidence": best_total["confidence"],
            "total_score": best_total["total_score"],
            "source": "enhanced_total_detection",
            "related_label": best_total.get("related_label", "")
        }
        
        return result
    
    # Se non abbiamo trovato nulla, proviamo con una ricerca basata sulle caratteristiche
    return find_totals_by_characteristics(ocr_result)

def find_totals_by_characteristics(ocr_result):
    """
    Cerca importi che hanno caratteristiche tipiche di un totale, anche in assenza di etichette.
    Caratteristiche: importo grande, posizione in basso a destra, numeri con decimali, presenza di valute.
    """
    print("Cercando importi con caratteristiche da totale (senza etichette)...")
    
    # Calcola dimensioni del documento
    doc_height = max([ocr_result["top"][i] + ocr_result["height"][i] for i in range(len(ocr_result["text"]))])
    doc_width = max([ocr_result["left"][i] + ocr_result["width"][i] for i in range(len(ocr_result["text"]))])
    
    potential_totals = []
    
    # Cerca importi nel documento
    for i, text in enumerate(ocr_result["text"]):
        if not is_amount_format(text):
            continue
            
        # Posizione relativa nel documento
        x_pos = (ocr_result["left"][i] + ocr_result["width"][i]/2) / doc_width if doc_width > 0 else 0.5
        y_pos = (ocr_result["top"][i] + ocr_result["height"][i]/2) / doc_height if doc_height > 0 else 0.5
        
        # Score per posizione (i totali tendono ad essere in basso a destra)
        position_score = 0
        if y_pos > 0.6:  # Nella metà inferiore
            position_score += 50 * (y_pos - 0.6) * 2.5  # Max 50 punti
        if x_pos > 0.5:  # Nella metà destra
            position_score += 30 * (x_pos - 0.5) * 2  # Max 30 punti
        
        # Confidenza OCR di base
        base_confidence = ocr_result["conf"][i]
        
        # Formato (decimali, valute)
        format_score = 0
        has_decimals = "," in text or "." in text
        if has_decimals:
            format_score += 20
        
        # Verifiche per valute (€, $, ecc.)
        surrounding_text = ""
        # Controlla token prima e dopo
        for offset in [-1, 1]:
            idx = i + offset
            if 0 <= idx < len(ocr_result["text"]):
                surrounding_text += ocr_result["text"][idx] + " "
        
        # Cerca simboli di valuta nel testo o nei token vicini
        currency_symbols = ["€", "$", "£", "¥", "EUR", "USD", "GBP", "JPY", "CHF", "CAD"]
        has_currency = any(symbol in text for symbol in currency_symbols) or any(symbol in surrounding_text for symbol in currency_symbols)
        if has_currency:
            format_score += 30
            
        # Calcola punteggio totale
        try:
            numeric_value = parse_amount_to_float(text)
            
            # Bonus per importi grandi
            size_score = min(50, numeric_value / 200)  # Max 50 punti
            
            total_score = base_confidence + position_score + format_score + size_score
            
            potential_totals.append({
                "text": text,
                "value": numeric_value,
                "confidence": base_confidence,
                "position_score": position_score,
                "format_score": format_score,
                "size_score": size_score,
                "total_score": total_score,
                "index": i,
                "has_decimals": has_decimals,
                "has_currency": has_currency,
                "x_position": x_pos,
                "y_position": y_pos
            })
            
            print(f"Potenziale TOTAL per caratteristiche: '{text}' "
                  f"(pos: {position_score:.1f}, formato: {format_score:.1f}, "
                  f"grandezza: {size_score:.1f}, totale: {total_score:.1f})")
        except ValueError:
            continue
    
    # Trova l'importo più grande
    max_value = 0
    if potential_totals:
        max_value = max(potential_totals, key=lambda x: x["value"])["value"]
        
        # Bonus per gli importi più grandi (più probabilmente totali)
        for total in potential_totals:
            # Se l'importo è almeno il 80% del massimo, probabilmente è un totale
            if total["value"] > max_value * 0.8:
                total["total_score"] += 40
                print(f"  Bonus per importo vicino al massimo: {total['text']} -> +40")
            # Penalizza gli importi molto piccoli rispetto al massimo
            elif total["value"] < max_value * 0.5:
                total["total_score"] -= 30
                print(f"  Penalità per importo piccolo: {total['text']} -> -30")
    
    # Seleziona il miglior candidato
    if potential_totals:
        # Ordina per punteggio
        potential_totals.sort(key=lambda x: x["total_score"], reverse=True)
        best_total = potential_totals[0]
        
        print(f"Miglior candidato TOTAL: '{best_total['text']}' (score: {best_total['total_score']:.1f})")
        
        # Crea il risultato nel formato atteso
        result = {
            "text": best_total["text"],
            "indices": [best_total["index"]],
            "confidence": best_total["confidence"],
            "total_score": best_total["total_score"],
            "source": "characteristics_based_detection"
        }
        
        return result
    
    return None

##############################################################################################################################################################################################

def is_valid_date_format(text):
    """
    Verifica se il testo è nel formato data valido.
    Supporta formati come:
    - DD/MM/YYYY (07/05/2025)
    - DD-MM-YYYY (07-05-2025)
    - DD.MM.YYYY (07.05.2025)
    - D Month YYYY (7 May 2025, 7 maggio 2025)
    - Month D, YYYY (May 7, 2025, maggio 7, 2025)
    - YYYY-MM-DD (2025-05-07, formato ISO)
    """
    import re
    from datetime import datetime

    text = text.strip()
    
    # Formati numerici separati da /, - o .
    if re.match(r'^(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})$', text):
        # Estrai i componenti della data
        parts = re.split(r'[/.-]', text)
        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        
        # Verifica che giorno e mese siano in range valido
        if not (1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100):
            return False
            
        # Verifica ulteriore per giorni validi nei mesi specifici
        days_in_month = [0, 31, 29 if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0 else 28, 
                         31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        if day > days_in_month[month]:
            return False
            
        return True
    
    # Formato ISO: YYYY-MM-DD
    if re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', text):
        parts = text.split('-')
        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
        
        if not (1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100):
            return False
            
        days_in_month = [0, 31, 29 if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0 else 28, 
                         31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        if day > days_in_month[month]:
            return False
            
        return True
    
    # Formati testuali: D Month YYYY o Month D, YYYY
    months_it = ['gennaio', 'febbraio', 'marzo', 'aprile', 'maggio', 'giugno', 
                'luglio', 'agosto', 'settembre', 'ottobre', 'novembre', 'dicembre']
    months_en = ['january', 'february', 'march', 'april', 'may', 'june', 
                'july', 'august', 'september', 'october', 'november', 'december']
    months_fr = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin', 
                'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
    
    months_short_it = [m[:3] for m in months_it]
    months_short_en = [m[:3] for m in months_en]
    months_short_fr = [m[:3] for m in months_fr]
    
    # Combina tutti i mesi
    all_months = months_it + months_en + months_fr + months_short_it + months_short_en + months_short_fr
    
    # Formato: D Month YYYY
    pattern1 = r'^(\d{1,2})\s+([a-zA-Z]+)\s+(\d{4})$'
    match1 = re.match(pattern1, text, re.IGNORECASE)
    if match1:
        day, month_name, year = match1.groups()
        day, year = int(day), int(year)
        month_name = month_name.lower()
        
        # Trova il numero del mese
        month = None
        for i, month_list in enumerate([months_it, months_en, months_fr, months_short_it, months_short_en, months_short_fr]):
            if month_name in [m.lower() for m in month_list]:
                month = (month_list.index(month_name) % 12) + 1
                break
        
        if month and 1 <= day <= 31 and 1900 <= year <= 2100:
            days_in_month = [0, 31, 29 if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0 else 28, 
                            31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            if day <= days_in_month[month]:
                return True
    
    # Formato: Month D, YYYY
    pattern2 = r'^([a-zA-Z]+)\s+(\d{1,2})(?:,|\.|\s+)?\s*(\d{4})$'
    match2 = re.match(pattern2, text, re.IGNORECASE)
    if match2:
        month_name, day, year = match2.groups()
        day, year = int(day), int(year)
        month_name = month_name.lower()
        
        month = None
        for i, month_list in enumerate([months_it, months_en, months_fr, months_short_it, months_short_en, months_short_fr]):
            if month_name in [m.lower() for m in month_list]:
                month = (month_list.index(month_name) % 12) + 1
                break
        
        if month and 1 <= day <= 31 and 1900 <= year <= 2100:
            days_in_month = [0, 31, 29 if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0 else 28, 
                            31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            if day <= days_in_month[month]:
                return True
    
    return False

def is_amount_format(text):
    """
    Verifica se il testo rappresenta un importo numerico valido.
    """
    import re
    
    # Rimuovi spazi iniziali e finali
    text = text.strip()
    if not text:
        return False
    
    # Rimuovi valute e simboli
    for symbol in ["€", "$", "£", "¥", "EUR", "USD", "GBP", "JPY"]:
        text = text.replace(symbol, "").strip()
    
    # Verifica formati numerici (supporta punti, virgole e spazi come separatori)
    # Formato europeo: 1.234,56 o 1234,56
    if re.match(r'^-?(\d{1,3}(?:[.\s]\d{3})*|\d+),\d{1,4}$', text):
        return True
    
    # Formato US: 1,234.56 o 1234.56
    if re.match(r'^-?(\d{1,3}(?:[,\s]\d{3})*|\d+)\.\d{1,4}$', text):
        return True
    
    # Numeri interi con separatori
    if re.match(r'^-?(\d{1,3}(?:[.,\s]\d{3})+|\d+)$', text):
        return True
    
    # Numeri semplici (almeno 1 cifra)
    if re.match(r'^-?\d+$', text):
        return True
    
    return False

def is_invoice_number_format(text):
    """
    Verifica se il testo ha un formato compatibile con un numero di fattura.
    Supporta formati come:
    - 2023/001
    - 001-2023
    - F123456
    - INV-001-2023
    - A/123/23
    - 00123 (numeri lunghi almeno 4-5 cifre)
    """
    import re
    
    text = str(text).strip()
    if not text:
        return False
    
    # Pattern comuni per numeri di fattura
    patterns = [
        # Anno/Numero (2023/001)
        r'^(20\d{2})[/.-](\d{2,6})$',
        
        # Numero/Anno (001/2023)
        r'^(\d{2,6})[/.-](20\d{2})$',
        
        # Prefisso-Numero (F-12345, INV-12345)
        r'^([a-zA-Z]{1,4})[/.-]?(\d{3,8})$',
        
        # Numero con prefisso e anno (INV-001-2023, F/001/23)
        r'^([a-zA-Z]{1,4})[/.-]?(\d{2,6})[/.-](20\d{2}|2[0-9]|[0-9]{2})$',
        
        # Solo numeri, ma almeno 4 cifre 
        r'^(\d{4,10})$',
        
        # Formato con lettere e numeri separati (A/123/23, INV/2023/001)
        r'^([a-zA-Z]{1,4})[/.-](\d{1,6})[/.-](\d{2,4})$'
    ]
    
    for pattern in patterns:
        if re.match(pattern, text):
            return True
            
    # Evita di interpretare date come numeri di fattura
    if is_valid_date_format(text):
        return False
        
    # Evita di interpretare importi come numeri di fattura
    if is_amount_format(text):
        return False
        
    return False

##############################################################################################################################################################################################

def export_to_excel(structured_data, output_path):
    """
    Versione migliorata della funzione per generare Excel.
    Formatta meglio il file e aggiunge validazioni.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        print(f"Creazione file Excel: {output_path}")
        
        # Inizializza il file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Dati Estratti"
        
        # Stili per l'intestazione
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        centered = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Creare un DataFrame vuoto con le colonne per tutte le entità
        entity_types = ["VENDOR", "CUSTOMER", "DATE", "TOTAL", "INVOICE_NUMBER"]
        
        # Inizializza liste per ogni tipo di entità
        entity_data = {entity_type: [] for entity_type in entity_types}
        
        # Raccogliere tutti i dati dalle entità trovate
        for entity_type, entries in structured_data.items():
            if entity_type in entity_types and entries:
                # Prendi solo il testo dell'entità con la confidenza più alta per ogni tipo
                best_entry = max(entries, key=lambda x: x.get('total_score', x.get('confidence', 0)))
                
                # Pulizia del testo (rimuove spazi extra, caratteri di controllo, ecc.)
                clean_text = best_entry['text'].strip()
                
                # Validazione specifica per tipo
                if entity_type == "TOTAL" and is_amount_format(clean_text):
                    # Per i totali, assicurati che il formato sia corretto
                    entity_data[entity_type].append(clean_text)
                elif entity_type == "DATE" and is_valid_date_format(clean_text):
                    # Per le date, assicurati che il formato sia corretto
                    entity_data[entity_type].append(clean_text)
                else:
                    entity_data[entity_type].append(clean_text)
            else:
                # Se il tipo di entità non esiste nei dati strutturati
                if entity_type in entity_types:
                    entity_data[entity_type].append("")
        
        # Riempi eventuali liste vuote
        for entity_type in entity_types:
            if not entity_data[entity_type]:
                entity_data[entity_type].append("")
        
        # Crea un DataFrame con una sola riga contenente i dati migliori per ogni entità
        df = pd.DataFrame({entity_type: entity_data[entity_type][:1] for entity_type in entity_types})
        
        # Traduci i nomi delle colonne in italiano
        column_translations = {
            "VENDOR": "FORNITORE",
            "CUSTOMER": "CLIENTE",
            "DATE": "DATA",
            "TOTAL": "TOTALE",
            "INVOICE_NUMBER": "NUMERO FATTURA"
        }
        
        # Aggiungi intestazioni
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_translations.get(column, column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border
            
            # Adatta la larghezza della colonna
            ws.column_dimensions[cell.column_letter].width = max(20, len(cell.value) + 5)
        
        # Aggiungi dati
        row_idx = 2
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical='center')
                cell.border = border
            row_idx += 1
        
        # Salva il file Excel
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            print(f"✅ File Excel salvato con successo in: {output_path}")
            return True
        except Exception as e_save:
            print(f"⚠️ Errore nel salvare il file Excel: {e_save}")
            return False
            
    except ImportError as e_imp:
        print(f"⚠️ Impossibile creare il file Excel. Libreria mancante: {e_imp}")
        print("Installa le librerie necessarie con: pip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"⚠️ Errore nella creazione del file Excel: {e}")
        traceback.print_exc()
        return False

##############################################################################################################################################################################################

def main():
    import argparse
    from datetime import datetime
    parser = argparse.ArgumentParser(description="Inferenza con modello LayoutLM per analisi fatture")
    parser.add_argument("--pdf", required=True, help="Percorso al file PDF da analizzare")
    parser.add_argument("--model", help="Directory del modello LayoutLM addestrato")
    parser.add_argument("--output", help="Directory di output per i risultati")
    parser.add_argument("--poppler", help="Percorso alla directory Poppler")
    parser.add_argument("--tesseract", help="Percorso alleseguibile Tesseract")
    parser.add_argument("--debug", action="store_true", help="Abilita modalità debug")

    args = parser.parse_args()
    start_time = datetime.now()
    print(f"Inizio esecuzione inferenza: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

    model_dir = Path(args.model or CONFIG["MODEL_DIR"])
    output_base_dir = Path(args.output or CONFIG["OUTPUT_DIR"])
    poppler_path = args.poppler or CONFIG["POPPLER_PATH"]
    tesseract_path = args.tesseract or CONFIG["TESSERACT_PATH"]

    if args.debug:
        print("\n=== MODALITÀ DEBUG ATTIVA ===")
        print(f"Modello: {model_dir}")
        print(f"Output: {output_base_dir}")
        print(f"Poppler: {poppler_path}")
        print(f"Tesseract: {tesseract_path}")

    try:
        print(f"Caricamento del modello da: {model_dir}")
        if not model_dir.exists() or not (model_dir / "pytorch_model.bin").exists() or not (model_dir / "config.json").exists():
            raise FileNotFoundError(f"Directory del modello \t{model_dir}\t non valida o file mancanti.")
        
        model_config = LayoutLMConfig.from_pretrained(str(model_dir), local_files_only=True)
        model = LayoutLMForTokenClassification.from_pretrained(str(model_dir), config=model_config, local_files_only=True)
        tokenizer = load_local_tokenizer(str(model_dir))
        model.eval()
        print("Modello e tokenizer caricati con successo.")

        pdf_file = Path(args.pdf)
        if not pdf_file.exists():
            raise FileNotFoundError(f"File PDF non trovato: {pdf_file}")

        images = convert_pdf_to_images(pdf_file, poppler_path=poppler_path)
        if not images:
            raise ValueError("Nessuna immagine estratta dal PDF.")

        all_results = []
        pdf_name = pdf_file.stem

        for page_idx, image in enumerate(images):
            print(f"\n--- Elaborazione pagina {page_idx + 1} --- ")
            page_output_dir = output_base_dir / pdf_name / f"page_{page_idx + 1}"
            os.makedirs(page_output_dir, exist_ok=True)

            words, boxes, word_indices_ocr, ocr_full_result = process_image_with_ocr(image, tesseract_path=tesseract_path)
            if not words:
                print(f"Nessuna parola trovata tramite OCR per la pagina {page_idx + 1}. Salto.")
                continue

            normalized_boxes_for_tokenizer = [normalize_box(b, image.width, image.height) for b in boxes]

            print(f"Tokenizzazione di {len(words)} parole...")
            try:
                encoding = tokenizer(
                    words,
                    boxes=normalized_boxes_for_tokenizer,
                    padding="max_length",
                    truncation=True,
                    max_length=CONFIG["MAX_SEQ_LENGTH"],
                    return_tensors="pt",
                    is_split_into_words=True
                )
            except TypeError:
                 print("Fallback: tokenizzazione senza argomento boxes.")
                 encoding = tokenizer(
                    words,
                    padding="max_length",
                    truncation=True,
                    max_length=CONFIG["MAX_SEQ_LENGTH"],
                    return_tensors="pt",
                    is_split_into_words=True
                )

            if "bbox" not in encoding:
                 print("Aggiunta manuale di bbox all'encoding.")
                 aligned_boxes = []
                 word_ids_enc = encoding.word_ids()
                 for i, token_id_enc in enumerate(encoding["input_ids"][0]):
                     if token_id_enc == tokenizer.cls_token_id or token_id_enc == tokenizer.sep_token_id or token_id_enc == tokenizer.pad_token_id:
                         aligned_boxes.append([0, 0, 0, 0])
                     else:
                         word_idx_enc = word_ids_enc[i]
                         if word_idx_enc is not None and word_idx_enc < len(normalized_boxes_for_tokenizer):
                             aligned_boxes.append(normalized_boxes_for_tokenizer[word_idx_enc])
                         else:
                             aligned_boxes.append([0,0,0,0])
                 if len(aligned_boxes) < CONFIG["MAX_SEQ_LENGTH"]:
                     aligned_boxes.extend([[0,0,0,0]] * (CONFIG["MAX_SEQ_LENGTH"] - len(aligned_boxes)))
                 encoding["bbox"] = torch.tensor([aligned_boxes[:CONFIG["MAX_SEQ_LENGTH"]]], dtype=torch.long)

            input_ids = encoding["input_ids"]
            attention_mask = encoding["attention_mask"]
            token_type_ids = encoding["token_type_ids"]
            bbox = encoding["bbox"]

            with torch.no_grad():
                outputs = model(input_ids=input_ids, attention_mask=attention_mask, token_type_ids=token_type_ids, bbox=bbox)
            
            predictions_logits = outputs.logits.argmax(-1).squeeze().tolist()
            
            token_predictions = {}
            # Usa word_ids() per mappare i token alle parole originali dell'OCR
            word_ids_map = encoding.word_ids() # Ottieni la mappa parola-token

            for token_idx, pred_id in enumerate(predictions_logits):
                word_idx_from_map = word_ids_map[token_idx]
                if word_idx_from_map is not None: # Solo token che corrispondono a parole
                    original_word_ocr_index = word_indices_ocr[word_idx_from_map] # Mappa all'indice OCR originale
                    label = id2label[pred_id]
                    
                    # Aggrega le predizioni per parola (se una parola è divisa in più token)
                    if str(original_word_ocr_index) not in token_predictions:
                        token_predictions[str(original_word_ocr_index)] = label
            
            print(f"Numero di predizioni a livello di token mappate a parole OCR: {len(token_predictions)}")

            # 1. Estrai i dati strutturati con il metodo standard
            structured_data = extract_structured_data(ocr_full_result, token_predictions)
            
            # 2. Applica le euristiche di posizione
            structured_data = apply_positional_heuristics(structured_data, ocr_full_result)
            
            # 3. Cerca specificamente i campi TOTAL con maggiore precisione
            total_result = extract_total_with_enhanced_detection(ocr_full_result)
            if total_result:
                if "TOTAL" not in structured_data:
                    structured_data["TOTAL"] = []
                structured_data["TOTAL"] = [total_result]
                print(f"Sostituito TOTAL con il valore trovato: {total_result['text']}")
            
            # 4. Cerca coppie etichetta-valore direttamente nel documento
            enhanced_data = enhance_extraction_with_direct_label_matching(structured_data, ocr_full_result)
            
            # 5. Applica regole specifiche come ultima risorsa
            final_data = add_rule_based_entities(ocr_full_result, enhanced_data)
            
            all_results.append({f"page_{page_idx + 1}": final_data})
            
            # Genera Excel per questa pagina
            excel_path = page_output_dir / f"{pdf_name}_page_{page_idx + 1}_structured.xlsx"
            export_to_excel(final_data, str(excel_path))
        
        # Genera Excel principale con i dati della prima pagina
        if all_results and f"page_1" in all_results[0]:
            excel_main_path = output_base_dir / f"{pdf_name}_data.xlsx"
            success = export_to_excel(all_results[0]["page_1"], str(excel_main_path))
            
            if success:
                print(f"\n✅ PROCESSO COMPLETATO: File Excel generato in: {excel_main_path}")
            else:
                print("\n⚠️ Errore nella generazione del file Excel principale")

    except Exception as e:
        print(f"Errore durante l'inferenza: {e}")
        traceback.print_exc()
    finally:
        end_time = datetime.now()
        print(f"Fine esecuzione inferenza: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Tempo totale di esecuzione: {end_time - start_time}")

if __name__ == "__main__":
    warnings.filterwarnings("ignore", category=UserWarning, module="tokenizers")
    main()



