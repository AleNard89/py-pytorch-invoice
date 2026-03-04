import os
import traceback
from PIL import Image, ImageDraw, ImageFont
from utils import is_valid_date_format, is_amount_format


def visualize_predictions(image, ocr_result, predictions, output_path):
    result_image = image.copy()
    draw = ImageDraw.Draw(result_image)
    colors = {
        "VENDOR": (255,0,0,150), "CUSTOMER": (0,0,255,150), "DATE": (0,255,0,150),
        "TOTAL": (128,0,128,150), "ITEM": (255,165,0,150), "QUANTITY": (0,255,255,150),
        "PRICE": (255,0,255,150), "INVOICE_NUMBER": (165,42,42,150), "O": (200,200,200,50)
    }
    legend_items = []
    entity_boxes = {}
    current_entity, current_text, current_box, current_indices = None, [], [float("inf"), float("inf"), 0, 0], []
    
    sorted_indices = sorted(predictions.keys(), key=lambda x: (ocr_result["top"][int(x)], ocr_result["left"][int(x)]))
    
    for word_idx in sorted_indices:
        i = int(word_idx)
        word = ocr_result["text"][i].strip()
        if not word: continue
        label = predictions[word_idx]
        x, y, w, h = ocr_result["left"][i], ocr_result["top"][i], ocr_result["width"][i], ocr_result["height"][i]

        if label.startswith("B-"):
            if current_entity and current_text:
                entity_type = current_entity
                if entity_type not in entity_boxes: entity_boxes[entity_type] = []
                avg_conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                entity_boxes[entity_type].append((" ".join(current_text), current_box, avg_conf))
                if entity_type not in [item[0] for item in legend_items]:
                    legend_items.append((entity_type, colors.get(entity_type, (0,0,0,150))))
            current_entity = label[2:]
            current_text = [word]
            current_box = [x, y, x + w, y + h]
            current_indices = [i]
        elif label.startswith("I-") and current_entity == label[2:]:
            current_text.append(word)
            current_box[0] = min(current_box[0], x)
            current_box[1] = min(current_box[1], y)
            current_box[2] = max(current_box[2], x + w)
            current_box[3] = max(current_box[3], y + h)
            current_indices.append(i)
        elif label == "O":
            if current_entity and current_text:
                entity_type = current_entity
                if entity_type not in entity_boxes: entity_boxes[entity_type] = []
                avg_conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
                entity_boxes[entity_type].append((" ".join(current_text), current_box, avg_conf))
                if entity_type not in [item[0] for item in legend_items]:
                    legend_items.append((entity_type, colors.get(entity_type, (0,0,0,150))))
            current_entity, current_text, current_box, current_indices = None, [], [float("inf"), float("inf"), 0, 0], []
            
    if current_entity and current_text:
        entity_type = current_entity
        if entity_type not in entity_boxes: entity_boxes[entity_type] = []
        avg_conf = sum(ocr_result["conf"][j] for j in current_indices) / len(current_indices) if current_indices else 0
        entity_boxes[entity_type].append((" ".join(current_text), current_box, avg_conf))
        if entity_type not in [item[0] for item in legend_items]:
            legend_items.append((entity_type, colors.get(entity_type, (0,0,0,150))))

    for entity_type, entries in entity_boxes.items():
        color = colors.get(entity_type, (0,0,0,150))
        for text, box, confidence in entries:
            x1, y1, x2, y2 = box
            overlay = Image.new("RGBA", result_image.size, (0,0,0,0))
            draw_overlay = ImageDraw.Draw(overlay)
            draw_overlay.rectangle([x1, y1, x2, y2], fill=color)
            if result_image.mode != "RGBA": result_image = result_image.convert("RGBA")
            result_image = Image.alpha_composite(result_image, overlay)
            draw = ImageDraw.Draw(result_image)
            draw.rectangle([x1, y1, x2, y2], outline=color[:3], width=2)
            try:
                font = ImageFont.truetype("arial.ttf", 12)
            except IOError:
                font = ImageFont.load_default()
            label_text = f"{entity_type}: {text[:30]}{'...' if len(text) > 30 else ''}"

            try: label_w, label_h = draw.textbbox((0,0), label_text, font=font)[2:]
            except AttributeError: label_w, label_h = draw.textsize(label_text, font=font)
            draw.rectangle([x1, y1 - label_h - 4, x1 + label_w + 4, y1], fill=color[:3])
            draw.text((x1 + 2, y1 - label_h - 2), label_text, fill=(255,255,255), font=font)

    if legend_items:
        legend_x, legend_y, legend_width = 10, 10, 150
        legend_height = len(legend_items) * 20 + 30
        draw.rectangle([legend_x, legend_y, legend_x + legend_width, legend_y + legend_height], fill=(0,0,0,128), outline=(255,255,255))
        try: title_font = ImageFont.truetype("arial.ttf", 14)
        except IOError: title_font = ImageFont.load_default()
        draw.text((legend_x + 5, legend_y + 5), "Entità rilevate:", fill=(255,255,255), font=title_font)
        for i, (entity_type, color) in enumerate(legend_items):
            item_y = legend_y + 25 + i * 20
            draw.rectangle([legend_x + 5, item_y, legend_x + 15, item_y + 10], fill=color[:3], outline=(255,255,255))
            draw.text((legend_x + 20, item_y - 2), entity_type, fill=(255,255,255), font=font)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    result_image = result_image.convert("RGB")
    result_image.save(output_path)
    print(f"Visualizzazione salvata in: {output_path}")
    return result_image


def export_to_excel(structured_data, output_path):
    """
    Versione migliorata della funzione per generare Excel.
    Formatta meglio il file e aggiunge validazioni.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        print(f"Creazione file Excel: {output_path}")
        
        # Inizializza il file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Dati Estratti"
        
        # Stili per l'intestazione
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        centered = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Creare un DataFrame vuoto con le colonne per tutte le entità
        entity_types = ["VENDOR", "CUSTOMER", "DATE", "TOTAL", "INVOICE_NUMBER"]
        
        # Inizializza liste per ogni tipo di entità
        entity_data = {entity_type: [] for entity_type in entity_types}
        
        # Raccogliere tutti i dati dalle entità trovate
        for entity_type, entries in structured_data.items():
            if entity_type in entity_types and entries:
                # Prendi solo il testo dell'entità con la confidenza più alta per ogni tipo
                best_entry = max(entries, key=lambda x: x.get('total_score', x.get('confidence', 0)))
                
                # Pulizia del testo (rimuove spazi extra, caratteri di controllo, ecc.)
                clean_text = best_entry['text'].strip()
                
                # Validazione specifica per tipo
                if entity_type == "TOTAL" and is_amount_format(clean_text):
                    # Per i totali, assicurati che il formato sia corretto
                    entity_data[entity_type].append(clean_text)
                elif entity_type == "DATE" and is_valid_date_format(clean_text):
                    # Per le date, assicurati che il formato sia corretto
                    entity_data[entity_type].append(clean_text)
                else:
                    entity_data[entity_type].append(clean_text)
            else:
                # Se il tipo di entità non esiste nei dati strutturati
                if entity_type in entity_types:
                    entity_data[entity_type].append("")
        
        # Riempi eventuali liste vuote
        for entity_type in entity_types:
            if not entity_data[entity_type]:
                entity_data[entity_type].append("")
        
        # Crea un DataFrame con una sola riga contenente i dati migliori per ogni entità
        df = pd.DataFrame({entity_type: entity_data[entity_type][:1] for entity_type in entity_types})
        
        # Traduci i nomi delle colonne in italiano
        column_translations = {
            "VENDOR": "FORNITORE",
            "CUSTOMER": "CLIENTE",
            "DATE": "DATA",
            "TOTAL": "TOTALE",
            "INVOICE_NUMBER": "NUMERO FATTURA"
        }
        
        # Aggiungi intestazioni
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_translations.get(column, column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = border
            
            # Adatta la larghezza della colonna
            ws.column_dimensions[cell.column_letter].width = max(20, len(cell.value) + 5)
        
        # Aggiungi dati
        row_idx = 2
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical='center')
                cell.border = border
            row_idx += 1
        
        # Salva il file Excel
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            print(f"✅ File Excel salvato con successo in: {output_path}")
            return True
        except Exception as e_save:
            print(f"⚠️ Errore nel salvare il file Excel: {e_save}")
            return False
            
    except ImportError as e_imp:
        print(f"⚠️ Impossibile creare il file Excel. Libreria mancante: {e_imp}")
        print("Installa le librerie necessarie con: pip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"⚠️ Errore nella creazione del file Excel: {e}")
        traceback.print_exc()
        return False
